from flask import Blueprint, render_template, request, redirect, jsonify, session, send_file
import io
import openpyxl
from openpyxl.styles import PatternFill, Font

from routes.admin import requiere_admin
from database import conectar

from services.comercial_service import (
    get_all_rutas,
    get_ruta,
    crear_ruta,
    ruta_existe,
    actualizar_tarifa_km_ruta,
    get_camioneros_por_ruta,
    get_all_camioneros_activos,
    asignar_camionero_a_ruta,
    desasociar_camionero_de_ruta,
    get_all_tipos_vehiculo,
    crear_tipo_vehiculo,
    get_all_tarifas,
    crear_tarifa,
    cotizar,
    guardar_cotizacion,
    get_all_cotizaciones,
    get_cotizacion_detalle,
    convertir_cotizacion_en_viaje
)

comercial_bp = Blueprint("comercial", __name__)


def get_viaje_id_por_cotizacion(cotizacion_id):
    conexion = conectar()
    cursor = conexion.cursor()

    cursor.execute("""
        SELECT id
        FROM viajes
        WHERE cotizacion_id = ?
        ORDER BY id DESC
        LIMIT 1
    """, (cotizacion_id,))

    viaje = cursor.fetchone()
    conexion.close()

    return viaje["id"] if viaje else None


@comercial_bp.route("/admin/comercial/rutas")
def rutas():
    if not requiere_admin():
        return redirect("/login")
    all_rutas = get_all_rutas()
    todos_camioneros = get_all_camioneros_activos()
    camioneros_por_ruta = {r["id"]: get_camioneros_por_ruta(r["id"]) for r in all_rutas}
    ids_asignados_por_ruta = {
        r["id"]: {c["id"] for c in camioneros_por_ruta[r["id"]]}
        for r in all_rutas
    }
    return render_template(
        "admin/comercial/rutas.html",
        rutas=all_rutas,
        todos_camioneros=todos_camioneros,
        camioneros_por_ruta=camioneros_por_ruta,
        ids_asignados_por_ruta=ids_asignados_por_ruta,
    )


@comercial_bp.route("/admin/comercial/rutas/nueva", methods=["POST"])
def nueva_ruta():
    if not requiere_admin():
        return redirect("/login")

    origen = request.form["origen"]
    destino = request.form["destino"]
    zona = request.form.get("zona", "")
    km = request.form["km"]

    if not ruta_existe(origen, destino):
        crear_ruta(origen, destino, zona, km)

    return redirect("/admin/comercial/rutas")


@comercial_bp.route("/admin/comercial/rutas/<int:ruta_id>/tarifa", methods=["POST"])
def actualizar_tarifa_ruta(ruta_id):
    if not requiere_admin():
        return redirect("/login")
    tarifa_km = request.form.get("tarifa_km", "").strip()
    actualizar_tarifa_km_ruta(ruta_id, tarifa_km if tarifa_km else None)
    return redirect("/admin/comercial/rutas")


@comercial_bp.route("/admin/comercial/rutas/<int:ruta_id>/editar", methods=["POST"])
def editar_ruta(ruta_id):
    if not requiere_admin():
        return redirect("/login")
    origen = request.form["origen"].strip()
    destino = request.form["destino"].strip()
    zona = request.form.get("zona", "").strip()
    km_oficiales = request.form.get("km_oficiales", "").strip()
    con = conectar()
    cur = con.cursor()
    cur.execute(
        "UPDATE rutas SET origen=?, destino=?, zona=?, km_oficiales=? WHERE id=?",
        (origen, destino, zona or None, km_oficiales or None, ruta_id)
    )
    con.commit()
    con.close()
    return redirect("/admin/comercial/rutas")


@comercial_bp.route("/admin/comercial/rutas/<int:ruta_id>/camioneros/asignar", methods=["POST"])
def asignar_camionero_ruta(ruta_id):
    if not requiere_admin():
        return redirect("/login")
    camionero_id = request.form.get("camionero_id")
    if camionero_id:
        asignar_camionero_a_ruta(ruta_id, int(camionero_id))
    return redirect("/admin/comercial/rutas")


@comercial_bp.route("/admin/comercial/rutas/<int:ruta_id>/camioneros/<int:camionero_id>/desasociar", methods=["POST"])
def desasociar_camionero_ruta(ruta_id, camionero_id):
    if not requiere_admin():
        return redirect("/login")
    desasociar_camionero_de_ruta(ruta_id, camionero_id)
    return redirect("/admin/comercial/rutas")


def _solo_admin():
    return "usuario" in session and session.get("rol") in ["admin", "operador"]


@comercial_bp.route("/admin/comercial/vehiculos")
def tipos_vehiculo():
    if not requiere_admin():
        return redirect("/login")
    if not _solo_admin():
        return redirect("/admin?access_error=Sin+permisos+para+acceder+a+Tipos+de+vehículo")

    return render_template(
        "admin/comercial/tipos_vehiculo.html",
        tipos=get_all_tipos_vehiculo()
    )


@comercial_bp.route("/admin/comercial/vehiculos/nuevo", methods=["POST"])
def nuevo_tipo_vehiculo():
    if not requiere_admin():
        return redirect("/login")
    if not _solo_admin():
        return redirect("/admin?access_error=Sin+permisos")

    nombre = request.form["nombre"]
    descripcion = request.form.get("descripcion", "")
    capacidad = request.form.get("capacidad_ton", "")

    crear_tipo_vehiculo(nombre, descripcion, capacidad)

    return redirect("/admin/comercial/vehiculos")


@comercial_bp.route("/admin/comercial/tarifas")
def tarifas():
    if not requiere_admin():
        return redirect("/login")
    if not _solo_admin():
        return redirect("/admin?access_error=Sin+permisos+para+acceder+a+Tarifas")

    return render_template(
        "admin/comercial/tarifas.html",
        tarifas=get_all_tarifas(),
        rutas=get_all_rutas(),
        tipos=get_all_tipos_vehiculo()
    )


@comercial_bp.route("/admin/comercial/tarifas/nueva", methods=["POST"])
def nueva_tarifa():
    if not requiere_admin():
        return redirect("/login")
    if not _solo_admin():
        return redirect("/admin?access_error=Sin+permisos")

    crear_tarifa(
        ruta_id=request.form["ruta_id"],
        tipo_vehiculo_id=request.form["tipo_vehiculo_id"],
        precio_cliente=request.form["precio_cliente"],
        pago_camionero=request.form["pago_camionero"],
        combustible_estimado=request.form.get("combustible_estimado", ""),
        vigencia_desde=request.form.get("vigencia_desde", ""),
        vigencia_hasta=request.form.get("vigencia_hasta", ""),
    )

    return redirect("/admin/comercial/tarifas")


@comercial_bp.route("/admin/comercial/cotizar")
def cotizar_view():
    if not requiere_admin():
        return redirect("/login")

    from routes.clientes import get_all_clientes
    from services.finanzas_service import get_configuracion

    cfg = get_configuracion()
    return render_template(
        "admin/comercial/cotizar.html",
        rutas=get_all_rutas(),
        tipos=get_all_tipos_vehiculo(),
        clientes=get_all_clientes(),
        tarifa_km_global=cfg.get("tarifa_km", 1.5),
    )


@comercial_bp.route("/admin/comercial/cotizar/calcular", methods=["POST"])
def calcular_cotizacion():
    if not requiere_admin():
        return redirect("/login")

    ruta_id = request.form["ruta_id"]
    tipo_vehiculo_id = request.form["tipo_vehiculo_id"]

    resultado = cotizar(ruta_id, tipo_vehiculo_id)

    if not resultado:
        return jsonify({"error": "No existe tarifa para esa combinación"}), 404

    return jsonify(resultado)


@comercial_bp.route("/admin/comercial/cotizar/guardar", methods=["POST"])
def guardar_cotizacion_view():
    if not requiere_admin():
        return redirect("/login")

    ruta_id = request.form["ruta_id"]
    tipo_vehiculo_id = request.form["tipo_vehiculo_id"]
    cliente_id = request.form.get("cliente_id")
    precio_override = request.form.get("precio_final")
    motivo = request.form.get("motivo")

    datos = cotizar(ruta_id, tipo_vehiculo_id)

    if not datos:
        return redirect("/admin/comercial/cotizar")

    guardar_cotizacion(
        datos,
        cliente_id if cliente_id else None,
        session["user_id"],
        precio_override,
        motivo
    )

    return redirect("/admin/comercial/cotizaciones")


@comercial_bp.route("/admin/comercial/cotizaciones")
def cotizaciones():
    if not requiere_admin():
        return redirect("/login")

    return render_template(
        "admin/comercial/cotizaciones.html",
        cotizaciones=get_all_cotizaciones()
    )


@comercial_bp.route("/admin/comercial/cotizaciones/plantilla")
def plantilla_cotizaciones():
    if not requiere_admin():
        return redirect("/login")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cotizaciones"

    headers = ["cliente", "ruta_id", "tipo_vehiculo_id", "precio_cliente", "pago_camionero", "observaciones"]
    fill = PatternFill(start_color="E86A2C", end_color="E86A2C", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = fill
        cell.font = font

    ws.append(["cliente@email.com", 1, 1, 500.00, 200.00, "Observaciones opcionales"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        as_attachment=True,
        download_name="plantilla_cotizaciones.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@comercial_bp.route("/admin/comercial/cotizaciones/importar", methods=["POST"])
def importar_cotizaciones():
    if not requiere_admin():
        return redirect("/login")

    archivo = request.files.get("archivo")
    if not archivo:
        return redirect("/admin/comercial/cotizaciones")

    wb = openpyxl.load_workbook(archivo)
    ws = wb.active

    con = conectar()
    cur = con.cursor()
    insertados = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        cliente_email, ruta_id, tipo_vehiculo_id, precio_cliente, pago_camionero, observaciones = (
            row[0], row[1], row[2], row[3], row[4], row[5] if len(row) > 5 else None
        )

        cliente_id = None
        if cliente_email:
            cur.execute("SELECT id FROM clientes WHERE email = ?", (str(cliente_email).strip(),))
            resultado = cur.fetchone()
            if resultado:
                cliente_id = resultado["id"]

        cur.execute("SELECT km_oficiales FROM rutas WHERE id = ?", (ruta_id,))
        ruta = cur.fetchone()
        if not ruta:
            continue

        km = ruta["km_oficiales"] or 0
        precio_cliente = float(precio_cliente or 0)
        pago_camionero_val = float(pago_camionero or 0)
        beneficio = round(precio_cliente - pago_camionero_val, 2)

        cur.execute("""
            INSERT INTO cotizaciones (
                cliente_id, ruta_id, tipo_vehiculo_id, km,
                precio_calculado, precio_final, pago_camionero,
                combustible_estimado, beneficio_estimado,
                modificado_manualmente, motivo_modificacion, estado
            ) VALUES (?, ?, ?, ?, ?, ?, ?, 0, ?, 0, ?, 'borrador')
        """, (
            cliente_id, ruta_id, tipo_vehiculo_id, km,
            precio_cliente, precio_cliente, pago_camionero_val,
            beneficio, str(observaciones) if observaciones else None
        ))
        insertados += 1

    con.commit()
    con.close()

    return redirect(f"/admin/comercial/cotizaciones?importado={insertados}")


@comercial_bp.route("/admin/comercial/cotizacion/<int:id>")
def ver_cotizacion(id):
    if not requiere_admin():
        return redirect("/login")

    cotizacion = get_cotizacion_detalle(id)

    if not cotizacion:
        return redirect("/admin/comercial/cotizaciones")

    viaje_id = get_viaje_id_por_cotizacion(id)

    return render_template(
        "admin/comercial/cotizacion_detalle.html",
        cotizacion=cotizacion,
        viaje_id=viaje_id
    )


@comercial_bp.route("/admin/comercial/cotizacion/<int:id>/convertir", methods=["POST"])
def convertir_cotizacion(id):
    if not requiere_admin():
        return redirect("/login")

    viaje_id = convertir_cotizacion_en_viaje(id)

    if not viaje_id:
        return redirect(f"/admin/comercial/cotizacion/{id}")

    return redirect(f"/admin/viajes/{viaje_id}/gestionar")