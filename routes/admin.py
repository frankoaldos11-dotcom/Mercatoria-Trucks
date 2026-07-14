import csv
import io
import json
import threading
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date, datetime
from urllib.parse import quote_plus

from flask import Blueprint, render_template, request, redirect, send_file, session, jsonify

from database import conectar, crear_checklist_viaje, INCIDENCIAS_CATEGORIAS, INCIDENCIAS_ESTADOS, sql_mes_actual
from db_config import USE_POSTGRES, ph
from extensions import bcrypt, mail
from flask_mail import Message
from services.comercial_service import (
    asignar_camionero_a_ruta, convertir_cotizacion_en_viaje, get_rutas_por_camionero,
)
from services.finanzas_service import calcular_liquidacion
from services.pdf_service import generar_factura_cliente, generar_pdf_orden_carga
from services.tramos_service import (
    ContinuidadError, completar_tramo, crear_tramos_viaje,
    obtener_tramos_viaje, tramos_completados, validar_continuidad,
)
from utils.constants import CAMIONERO_ESTADOS, VEHICULO_ESTADOS

admin_bp = Blueprint("admin", __name__, url_prefix="/admin")


def registrar_auditoria(accion, categoria, entidad=None, entidad_id=None, detalle=None):
    try:
        conexion = conectar()
        cursor = conexion.cursor()
        cursor.execute(f"""
            INSERT INTO auditoria (usuario, rol, accion, categoria, entidad, entidad_id, detalle)
            VALUES ({ph()}, {ph()}, {ph()}, {ph()}, {ph()}, {ph()}, {ph()})
        """, (
            session.get("usuario", "sistema"),
            session.get("rol", ""),
            accion, categoria, entidad, entidad_id, detalle
        ))
        conexion.commit()
        conexion.close()
    except Exception as e:
        from flask import current_app
        current_app.logger.error(f"AUDITORIA ERROR: {e}")


def _viaje_cerrado(viaje_id):
    con = conectar()
    cur = con.cursor()
    cur.execute(f"SELECT estado FROM viajes WHERE id = {ph()}", (viaje_id,))
    row = cur.fetchone()
    con.close()
    return bool(row) and (row["estado"] or "").lower() == "cerrado"


def _rutas_no_cubiertas(cursor, camionero_id, viaje_id, ruta_id_directo):
    """Devuelve la lista de rutas del viaje (dicts con ruta_id/origen/destino) que el
    camionero indicado NO tiene habilitadas en camionero_ruta. Lista vacía = las cubre todas.
    Si el viaje tiene tramos (multi-tramo) valida cada tramo en orden; si no, valida
    la ruta_id directa del viaje."""
    tramos = obtener_tramos_viaje(viaje_id)
    if tramos:
        rutas = [{"ruta_id": t["ruta_id"], "origen": t["origen"], "destino": t["destino"]} for t in tramos]
    elif ruta_id_directo:
        cursor.execute(f"SELECT id, origen, destino FROM rutas WHERE id = {ph()}", (ruta_id_directo,))
        r = cursor.fetchone()
        rutas = [{"ruta_id": r["id"], "origen": r["origen"], "destino": r["destino"]}] if r else []
    else:
        rutas = []

    no_cubiertas = []
    for ruta in rutas:
        cursor.execute(
            f"SELECT 1 FROM camionero_ruta WHERE camionero_id = {ph()} AND ruta_id = {ph()}",
            (camionero_id, ruta["ruta_id"])
        )
        if not cursor.fetchone():
            no_cubiertas.append(ruta)
    return no_cubiertas


def _registrar_historial(cursor, viaje_id, accion, detalle=""):
    """Inserta una fila en historial_viaje usando el cursor de la transacción ya
    abierta del endpoint llamador. No abre conexión propia ni comitea: el commit
    lo hace el endpoint como parte de su propia transacción."""
    cursor.execute(
        f"INSERT INTO historial_viaje (viaje_id, usuario, accion, detalle) "
        f"VALUES ({ph()}, {ph()}, {ph()}, {ph()})",
        (viaje_id, session.get("usuario", "sistema"), accion, detalle or None)
    )


def notificar_cliente_estado(viaje_id, nuevo_estado, email_cliente):
    if not email_cliente or "@" not in email_cliente:
        return
    try:
        from flask import current_app
        conexion = conectar()
        cursor = conexion.cursor()
        cursor.execute("SELECT clave, valor FROM configuracion_texto")
        txt = {r["clave"]: r["valor"] for r in cursor.fetchall()}
        conexion.close()
        if txt.get("mail_username"):
            current_app.config["MAIL_USERNAME"] = txt["mail_username"]
        if txt.get("mail_password"):
            current_app.config["MAIL_PASSWORD"] = txt["mail_password"]

        emojis = {
            "Asignado": "🚛", "En ruta": "📍", "Entregado": "✅",
            "Cancelado": "❌", "Carga recogida": "📦",
        }
        emoji = emojis.get(nuevo_estado, "📋")
        msg = Message(
            subject=f"{emoji} Tu envío #{viaje_id} está: {nuevo_estado} — Mercatoria Truck",
            recipients=[email_cliente],
        )
        msg.body = (
            f"Hola,\n\n"
            f"Tu envío #{viaje_id} ha cambiado de estado.\n\n"
            f"Estado actual: {nuevo_estado}\n\n"
            f"Puedes ver el detalle completo en tu portal:\n"
            f"https://mercatoria-trucks.onrender.com/cliente/viajes/{viaje_id}\n\n"
            f"— Mercatoria Truck\n"
        )
        mail.send(msg)
    except Exception as e:
        from flask import current_app
        current_app.logger.error(f"Error enviando email estado viaje {viaje_id}: {e}")


def requiere_admin():
    return "usuario" in session and session.get("rol") in ["admin", "operador"]


@admin_bp.route("/")
def dashboard():
    if not requiere_admin():
        return redirect("/login")

    conexion = conectar()
    cursor = conexion.cursor()

    cursor.execute("""
        SELECT
            COUNT(CASE WHEN LOWER(estado) IN ('pendiente','solicitado') THEN 1 END) AS pendientes,
            COUNT(CASE WHEN LOWER(estado) NOT IN ('entregado','cerrado','cancelado') THEN 1 END) AS en_curso,
            COUNT(CASE WHEN LOWER(estado) = 'entregado'                 THEN 1 END) AS entregados,
            COUNT(CASE WHEN LOWER(estado) = 'asignado'                  THEN 1 END) AS asignados,
            COUNT(CASE WHEN LOWER(estado) = 'cancelado'                 THEN 1 END) AS cancelados
        FROM viajes
    """)
    _est = cursor.fetchone()
    pendientes = _est["pendientes"]
    en_curso   = _est["en_curso"]
    entregados = _est["entregados"]
    asignados  = _est["asignados"]
    cancelados = _est["cancelados"]

    cursor.execute("SELECT COUNT(*) AS total FROM camioneros")
    camioneros = cursor.fetchone()["total"]

    cursor.execute("""
        SELECT
            COUNT(*) AS total,
            COUNT(CASE WHEN nombre IS NULL OR TRIM(nombre) = '' THEN 1 END) AS sin_nombre
        FROM clientes
    """)
    _cl = cursor.fetchone()
    clientes           = _cl["total"]
    clientes_sin_nombre = _cl["sin_nombre"]

    cursor.execute("""
        SELECT v.id,
               COALESCE(NULLIF(TRIM(cl.nombre), ''),
                        NULLIF(TRIM(u.nombre), ''),
                        v.cliente) AS cliente,
               v.origen, v.destino, v.estado
        FROM viajes v
        LEFT JOIN clientes cl ON v.cliente_id = cl.id
        LEFT JOIN usuarios u ON v.cliente = u.usuario
        WHERE LOWER(v.estado) IN ('pendiente', 'solicitado')
          AND v.deleted_at IS NULL
        ORDER BY v.id DESC
    """)
    lista = cursor.fetchall()

    # Ingresos y camionero top: solo para admin
    if session.get("rol") == "admin":
        cursor.execute(f"""
            SELECT COALESCE(SUM(COALESCE(NULLIF(precio_final,0), NULLIF(precio_cliente,0), NULLIF(precio,0), 0)), 0) AS total
            FROM viajes
            WHERE {sql_mes_actual()}
              AND LOWER(estado) != 'cancelado'
        """)
        ingresos_mes = round(cursor.fetchone()["total"], 2)

        cursor.execute("""
            SELECT camionero_nombre, COUNT(*) as total
            FROM viajes
            WHERE camionero_nombre IS NOT NULL AND camionero_nombre != ''
              AND LOWER(estado) != 'cancelado'
            GROUP BY camionero_nombre
            ORDER BY total DESC LIMIT 1
        """)
        row = cursor.fetchone()
        camionero_top = row["camionero_nombre"] if row else "—"

        cursor.execute("""
            SELECT COUNT(*) AS total FROM viajes
            WHERE LOWER(estado) IN ('entregado', 'cerrado')
              AND (estado_pago_camionero IS NULL OR estado_pago_camionero != 'Pagado')
              AND deleted_at IS NULL
        """)
        viajes_sin_pagar = cursor.fetchone()["total"]

        cursor.execute("""
            SELECT COUNT(*) AS total FROM viajes
            WHERE LOWER(estado) IN ('entregado', 'cerrado')
              AND fecha_cobro IS NULL
              AND deleted_at IS NULL
        """)
        viajes_sin_cobrar = cursor.fetchone()["total"]
    else:
        ingresos_mes = None
        camionero_top = None
        viajes_sin_pagar = None
        viajes_sin_cobrar = None

    # Solicitudes de eliminación pendientes (solo admin las ve)
    solicitudes_pendientes = []
    if session.get("rol") == "admin":
        try:
            cursor.execute("""
                SELECT id, entidad, entidad_id, nombre_entidad, solicitado_por, fecha_solicitud
                FROM solicitudes_eliminacion
                WHERE estado = 'Pendiente'
                ORDER BY fecha_solicitud DESC
            """)
            solicitudes_pendientes = cursor.fetchall()
        except Exception:
            pass

    cursor.execute("SELECT COUNT(*) AS total FROM incidencias WHERE estado = 'Abierta'")
    incidencias_abiertas = cursor.fetchone()["total"]

    cursor.execute("""
        SELECT i.id, i.viaje_id, i.categoria, i.descripcion, i.fecha_hora,
               v.origen, v.destino
        FROM incidencias i
        LEFT JOIN viajes v ON i.viaje_id = v.id
        WHERE i.estado = 'Abierta'
        ORDER BY i.fecha_hora DESC
        LIMIT 10
    """)
    incidencias_abiertas_lista = cursor.fetchall()

    conexion.close()

    return render_template(
        "admin/dashboard.html",
        pendientes=pendientes,
        en_curso=en_curso,
        entregados=entregados,
        asignados=asignados,
        cancelados=cancelados,
        camioneros=camioneros,
        clientes=clientes,
        lista=lista,
        ingresos_mes=ingresos_mes,
        camionero_top=camionero_top,
        clientes_sin_nombre=clientes_sin_nombre,
        solicitudes_pendientes=solicitudes_pendientes,
        incidencias_abiertas=incidencias_abiertas,
        incidencias_abiertas_lista=incidencias_abiertas_lista,
        viajes_sin_pagar=viajes_sin_pagar,
        viajes_sin_cobrar=viajes_sin_cobrar,
    )


@admin_bp.route("/viajes/nuevo", methods=["POST"])
def nuevo_viaje_admin():
    if not requiere_admin():
        return redirect("/login")

    ruta_ids = [r.strip() for r in request.form.getlist("ruta_id") if r.strip()]
    tipo_carga = request.form.get("tipo_carga", "").strip()
    tipo_transporte = request.form.get("tipo_transporte", "").strip()
    peso_str = request.form.get("peso_toneladas", "").strip()
    cantidad_str = request.form.get("cantidad_contenedores", "").strip()
    numero_contenedor = request.form.get("numero_contenedor", "").strip()
    notas = request.form.get("notas", "").strip()
    obs_operativas = request.form.get("observaciones_operativas", "").strip()
    cliente_id_str = request.form.get("cliente_id", "").strip()

    def _error_con_contexto(mensaje):
        con = conectar()
        cur = con.cursor()
        ctx = _contexto_lista_viajes(cur)
        con.close()
        return render_template("admin/viajes.html", error=mensaje, form_data=request.form, **ctx)

    if not ruta_ids or not tipo_carga:
        return _error_con_contexto("Selecciona la ruta y el tipo de carga")

    ruta_ids_int = [int(r) for r in ruta_ids]

    con = conectar()
    cur = con.cursor()

    placeholders = ",".join("?" for _ in ruta_ids_int)
    cur.execute(
        f"SELECT id, origen, destino FROM rutas WHERE id IN ({placeholders}) AND activa = 1",
        ruta_ids_int
    )
    rutas_por_id = {r["id"]: r for r in cur.fetchall()}
    if len(rutas_por_id) != len(set(ruta_ids_int)):
        con.close()
        return _error_con_contexto("Ruta no válida")
    rutas_ordenadas = [rutas_por_id[rid] for rid in ruta_ids_int]

    try:
        validar_continuidad(rutas_ordenadas)
    except ContinuidadError as e:
        con.close()
        return _error_con_contexto(str(e))

    ruta = rutas_ordenadas[0]
    ruta_id = ruta_ids_int[0]
    origen_viaje = rutas_ordenadas[0]["origen"]
    destino_viaje = rutas_ordenadas[-1]["destino"]

    cliente_id = int(cliente_id_str) if cliente_id_str else None
    cliente_nombre = ""
    if cliente_id:
        cur.execute("SELECT nombre, email FROM clientes WHERE id = ?", (cliente_id,))
        c = cur.fetchone()
        if c:
            cliente_nombre = c["nombre"] or c["email"] or ""

    peso_toneladas = None
    try:
        peso_toneladas = float(peso_str) if peso_str else None
    except ValueError:
        pass

    cantidad_contenedores = None
    if cantidad_str:
        try:
            cantidad_contenedores = int(cantidad_str)
        except ValueError:
            pass

    cur.execute("""
        INSERT INTO viajes (
            cliente, cliente_id, ruta_id, origen, destino,
            precio, combustible, comision, beneficio, estado,
            observaciones, prioridad,
            tipo_carga, tipo_transporte, cantidad_contenedores,
            numero_contenedor, peso_toneladas, observaciones_operativas
        )
        VALUES (?, ?, ?, ?, ?, 0, 0, 0, 0, 'Pendiente',
                ?, 'Normal', ?, ?, ?, ?, ?, ?)
    """, (
        cliente_nombre, cliente_id, ruta_id, origen_viaje, destino_viaje,
        notas or None,
        tipo_carga,
        tipo_transporte or None,
        cantidad_contenedores,
        numero_contenedor or None,
        peso_toneladas,
        obs_operativas or None,
    ))
    viaje_id = cur.lastrowid
    crear_checklist_viaje(cur, viaje_id)
    if len(ruta_ids_int) > 1:
        crear_tramos_viaje(cur, viaje_id, ruta_ids_int)
    con.commit()
    con.close()

    registrar_auditoria("creó viaje", "viajes", "viaje", viaje_id)
    return redirect(f"/admin/viajes/{viaje_id}/gestionar")


@admin_bp.route("/api/origenes-destinos")
def api_origenes_destinos():
    if not requiere_admin():
        return jsonify([])
    con = conectar()
    cur = con.cursor()
    cur.execute("SELECT DISTINCT origen FROM rutas WHERE activa = 1 AND origen IS NOT NULL ORDER BY origen")
    origenes = [r["origen"] for r in cur.fetchall()]
    cur.execute("SELECT DISTINCT destino FROM rutas WHERE activa = 1 AND destino IS NOT NULL ORDER BY destino")
    destinos = [r["destino"] for r in cur.fetchall()]
    con.close()
    return jsonify(sorted(set(origenes + destinos)))


def _contexto_lista_viajes(cursor):
    filtro = request.args.get("estado", "").strip().lower()
    buscar = request.args.get("buscar", "").strip()
    pagina = max(1, int(request.args.get("pagina", 1) or 1))
    por_pagina = 20

    condiciones = []
    params = []

    if filtro:
        condiciones.append(f"LOWER(v.estado) = {ph()}")
        params.append(filtro)

    if buscar:
        condiciones.append(f"""(
            COALESCE(c.nombre, v.cliente, '') LIKE {ph()}
            OR COALESCE(v.origen, '') LIKE {ph()}
            OR COALESCE(v.destino, '') LIKE {ph()}
            OR COALESCE(v.camionero_nombre, '') LIKE {ph()}
        )""")
        like = f"%{buscar}%"
        params.extend([like, like, like, like])

    where = ("WHERE " + " AND ".join(condiciones)) if condiciones else ""

    cursor.execute(f"""
        SELECT COUNT(*) AS total FROM viajes v
        {where}
    """, params)
    total = cursor.fetchone()["total"]

    total_paginas = max(1, (total + por_pagina - 1) // por_pagina)
    pagina = min(pagina, total_paginas)
    offset = (pagina - 1) * por_pagina

    cursor.execute(f"""
        SELECT v.id,
               COALESCE(c.nombre, v.cliente, 'Sin nombre') as cliente,
               v.origen, v.destino, v.estado, v.camionero_nombre,
               COALESCE(v.precio_final, v.precio_cliente, v.precio, 0) as precio,
               v.fecha_creacion,
               v.prioridad, v.tipo_carga,
               COALESCE(v.estado_pago_camionero, 'Pendiente') AS estado_pago_camionero,
               COALESCE(c.categoria, 'Normal') AS cliente_categoria,
               (SELECT COUNT(*) FROM incidencias i WHERE i.viaje_id = v.id AND i.estado != 'Resuelta') AS incidencias_abiertas
        FROM viajes v
        LEFT JOIN clientes c ON c.id = v.cliente_id
        {where}
        ORDER BY v.id DESC
        LIMIT {ph()} OFFSET {ph()}
    """, params + [por_pagina, offset])
    lista = cursor.fetchall()

    cursor.execute("SELECT id, nombre, origen, destino FROM rutas WHERE activa = 1 ORDER BY origen")
    rutas_list = cursor.fetchall()

    cursor.execute("SELECT id, nombre FROM clientes WHERE deleted_at IS NULL ORDER BY nombre")
    clientes_list = cursor.fetchall()

    return {
        "lista": lista,
        "filtro": filtro,
        "buscar": buscar,
        "pagina_actual": pagina,
        "total_paginas": total_paginas,
        "total": total,
        "rutas_list": rutas_list,
        "clientes_list": clientes_list,
    }


@admin_bp.route("/viajes")
def viajes():
    if not requiere_admin():
        return redirect("/login")

    conexion = conectar()
    cursor = conexion.cursor()
    ctx = _contexto_lista_viajes(cursor)
    conexion.close()

    return render_template("admin/viajes.html", **ctx)


def _parsear_observaciones(obs):
    """Extrae tipo_carga, peso y notas del texto de observaciones."""
    resultado = {"tipo_carga": "", "peso": "", "notas": "", "crudo": obs or ""}
    if not obs:
        return resultado
    for linea in obs.split("\n"):
        linea = linea.strip()
        if linea.startswith("Tipo de carga:"):
            partes = linea.split("|")
            resultado["tipo_carga"] = partes[0].replace("Tipo de carga:", "").strip()
            if len(partes) > 1:
                resultado["peso"] = partes[1].replace("Peso aprox.:", "").strip()
        elif linea.startswith("Notas:"):
            resultado["notas"] = linea.replace("Notas:", "").strip()
    return resultado


@admin_bp.route("/viaje/<int:id>")
@admin_bp.route("/viajes/<int:id>/gestionar")
def gestionar_viaje(id):
    if not requiere_admin():
        return redirect("/login")

    conexion = conectar()
    cursor = conexion.cursor()

    cursor.execute(f"""
        SELECT v.*,
               v.camionero_id as camionero_id
        FROM viajes v
        WHERE v.id = {ph()}
    """, (id,))
    viaje = cursor.fetchone()

    if not viaje:
        conexion.close()
        return redirect("/admin/viajes")

    cursor.execute("""
        SELECT c.id, c.nombre, c.telefono, c.estado,
               v.id       AS vehiculo_id,
               COALESCE(v.matricula, '') AS vehiculo_matricula,
               COALESCE(v.marca, '')    AS vehiculo_marca,
               COALESCE(v.modelo, '')   AS vehiculo_modelo
        FROM camioneros c
        INNER JOIN vehiculos v ON v.camionero_id = c.id AND v.activo = 1
        WHERE c.activo = 1 OR c.activo IS NULL
        ORDER BY c.nombre
    """)
    camioneros = cursor.fetchall()
    vehiculos = []

    tipo_vehiculo_nombre = None
    if viaje["tipo_vehiculo_id"]:
        cursor.execute(f"SELECT nombre FROM tipos_vehiculo WHERE id = {ph()}", (viaje["tipo_vehiculo_id"],))
        row = cursor.fetchone()
        tipo_vehiculo_nombre = row["nombre"] if row else None

    tarifa_info = None
    if viaje["tarifa_id"]:
        cursor.execute(f"""
            SELECT t.precio_cliente, tv.nombre AS tipo_nombre
            FROM tarifas t
            LEFT JOIN tipos_vehiculo tv ON t.tipo_vehiculo_id = tv.id
            WHERE t.id = {ph()}
        """, (viaje["tarifa_id"],))
        row = cursor.fetchone()
        if row:
            tarifa_info = f"#{viaje['tarifa_id']} — {row['tipo_nombre'] or 'N/A'} · ${row['precio_cliente']:.2f}/km"

    # Datos enriquecidos del cliente
    cliente_info = None
    if viaje["cliente_id"]:
        cursor.execute(
            f"SELECT nombre, telefono, empresa, email, COALESCE(categoria, 'Normal') AS categoria FROM clientes WHERE id = {ph()}",
            (viaje["cliente_id"],)
        )
        cliente_info = cursor.fetchone()

    # Nombre legible de la ruta y km oficiales
    ruta_display = None
    km_ruta = 0.0
    ids_en_ruta = []
    if viaje["ruta_id"]:
        cursor.execute(f"SELECT origen, destino, km_oficiales FROM rutas WHERE id = {ph()}", (viaje["ruta_id"],))
        ruta_row = cursor.fetchone()
        if ruta_row:
            ruta_display = f"{ruta_row['origen']} → {ruta_row['destino']}"
            km_ruta = float(ruta_row["km_oficiales"] or 0)
        cursor.execute(f"SELECT camionero_id FROM camionero_ruta WHERE ruta_id = {ph()}", (viaje["ruta_id"],))
        ids_en_ruta = [r["camionero_id"] for r in cursor.fetchall()]

    conexion.close()

    liquidacion = calcular_liquidacion(id)
    error = request.args.get("error")
    camionero_intentado = request.args.get("camionero_intentado")
    obs_parsed = _parsear_observaciones(viaje["observaciones"])

    _transiciones = {
        "solicitado":         ["Asignado", "En ruta", "Carga recogida", "Entregado", "Cancelado"],
        "pendiente":          ["Asignado", "En ruta", "Carga recogida", "Entregado", "Cancelado"],
        "asignado":           ["En ruta", "Carga recogida", "Entregado", "Cancelado"],
        "en ruta":            ["Carga recogida", "Entregado", "Asignado", "Cancelado"],
        "en_ruta":            ["Carga recogida", "Entregado", "Asignado", "Cancelado"],
        "carga recogida":     ["En ruta", "Entregado", "Cancelado"],
        "pendiente de pago":  ["Confirmado", "Entregado", "Cancelado"],
        "confirmado":         ["En ruta", "Carga recogida", "Entregado", "Cancelado"],
        "entregado":          ["En ruta", "Asignado", "Cancelado"],
        "cancelado":          ["Solicitado", "Asignado"],
        "cerrado":            [],
    }
    estado_norm = (viaje["estado"] or "").lower()
    estados_validos = _transiciones.get(estado_norm, ["Asignado", "En ruta", "Entregado", "Cancelado"])

    orden_faltantes = []
    if not viaje["cliente"]:
        orden_faltantes.append("cliente")
    if not viaje["camionero_nombre"] and not viaje["camionero_id"]:
        orden_faltantes.append("camionero")
    if not viaje["vehiculo_id"]:
        orden_faltantes.append("vehículo")
    if not viaje["origen"]:
        orden_faltantes.append("origen")
    if not viaje["destino"]:
        orden_faltantes.append("destino")
    orden_carga_ok = len(orden_faltantes) == 0
    orden_carga_tooltip = "Falta: " + ", ".join(orden_faltantes) if orden_faltantes else ""

    conexion2 = conectar()
    cursor2 = conexion2.cursor()
    cursor2.execute(f"""
        SELECT usuario, texto, fecha
        FROM notas_viaje
        WHERE viaje_id = {ph()}
        ORDER BY fecha DESC
    """, (id,))
    notas = cursor2.fetchall()

    # Checklist operativo — crear automáticamente si el viaje no lo tiene aún
    cursor2.execute(
        f"SELECT COUNT(*) AS cnt FROM viaje_checklist WHERE viaje_id = {ph()}", (id,)
    )
    if cursor2.fetchone()["cnt"] == 0:
        crear_checklist_viaje(cursor2, id)
        conexion2.commit()
    cursor2.execute(
        "SELECT id, item, completado, completado_por, fecha_completado "
        f"FROM viaje_checklist WHERE viaje_id = {ph()} ORDER BY id",
        (id,)
    )
    checklist = cursor2.fetchall()

    cursor2.execute(
        "SELECT id, categoria, descripcion, usuario, fecha_hora, estado "
        f"FROM incidencias WHERE viaje_id = {ph()} ORDER BY fecha_hora DESC",
        (id,)
    )
    incidencias = cursor2.fetchall()

    cursor2.execute(
        "SELECT usuario, accion, detalle, fecha_hora "
        f"FROM historial_viaje WHERE viaje_id = {ph()} ORDER BY fecha_hora DESC",
        (id,)
    )
    historial = cursor2.fetchall()
    conexion2.close()

    # Determine which checklist items are auto-completed based on current viaje state
    def _flt(row, *keys):
        for k in keys:
            try:
                v = row[k]
                if v is not None:
                    f = float(v)
                    if f > 0:
                        return f
            except (KeyError, IndexError, TypeError, ValueError):
                pass
        return 0.0

    _e = (viaje["estado"] or "").lower().strip()
    auto_completados = set()
    if _flt(viaje, "km", "kilometros") > 0:
        auto_completados.add("Km calculados")
    if _flt(viaje, "combustible", "combustible_estimado") > 0:
        auto_completados.add("Combustible calculado")
    if _flt(viaje, "pago_camionero", "camionero") > 0:
        auto_completados.add("Pago camionero calculado")
    if viaje["camionero_id"]:
        auto_completados.add("Camionero asignado")
    if viaje["vehiculo_id"]:
        auto_completados.add("Vehículo confirmado")
    if _e in {"carga recogida", "en ruta", "en_ruta", "entregado"}:
        auto_completados.add("Contenedor extraído / carga recogida")
    if _e == "entregado":
        auto_completados.add("Descarga realizada")
    try:
        if (viaje["estado_pago_camionero"] or "") == "Pagado":
            auto_completados.add("Pago camionero confirmado")
    except (KeyError, IndexError):
        pass

    checklist_map = {item["item"]: {"id": item["id"], "completado": bool(item["completado"])} for item in checklist}

    tramos = obtener_tramos_viaje(id)
    tramos_ok = tramos_completados(id)
    if tramos:
        km_ruta = sum(float(t["km_oficiales"] or 0) for t in tramos)
        ruta_display = " → ".join([tramos[0]["origen"]] + [t["destino"] for t in tramos])

    return render_template(
        "admin/gestionar_viaje.html",
        viaje=viaje,
        camioneros=camioneros,
        vehiculos=vehiculos,
        liquidacion=liquidacion,
        tipo_vehiculo_nombre=tipo_vehiculo_nombre,
        tarifa_info=tarifa_info,
        error=error,
        camionero_intentado=camionero_intentado,
        estados_validos=estados_validos,
        orden_carga_ok=orden_carga_ok,
        orden_carga_tooltip=orden_carga_tooltip,
        cliente_info=cliente_info,
        ruta_display=ruta_display,
        km_ruta=km_ruta,
        ids_en_ruta=ids_en_ruta,
        obs_parsed=obs_parsed,
        notas=notas,
        checklist=checklist,
        checklist_map=checklist_map,
        incidencias=incidencias,
        incidencias_categorias=INCIDENCIAS_CATEGORIAS,
        incidencias_estados=INCIDENCIAS_ESTADOS,
        auto_completados=auto_completados,
        historial=historial,
        tramos=tramos,
        tramos_ok=tramos_ok,
    )


@admin_bp.route("/viaje/<int:id>/tramo/<int:tramo_id>/completar", methods=["POST"])
def completar_tramo_admin(id, tramo_id):
    if not requiere_admin():
        return redirect("/login")
    if _viaje_cerrado(id):
        return redirect(f"/admin/viaje/{id}?error=El+viaje+está+cerrado+y+no+admite+cambios")

    con = conectar()
    cur = con.cursor()
    ok = completar_tramo(cur, id, tramo_id)
    if ok:
        _registrar_historial(cur, id, "Tramo completado", f"Tramo ID: {tramo_id}")
        con.commit()
        con.close()
        registrar_auditoria("completó tramo", "Viajes", "viaje", id, f"Tramo ID: {tramo_id}")
    else:
        con.close()
        return redirect(f"/admin/viajes/{id}/gestionar?error=Ese+tramo+no+se+puede+completar+todav%C3%ADa")

    return redirect(f"/admin/viajes/{id}/gestionar")


@admin_bp.route("/viaje/<int:viaje_id>/checklist/<int:item_id>/toggle", methods=["POST"])
def toggle_checklist(viaje_id, item_id):
    if not requiere_admin():
        return jsonify({"error": "no auth"}), 401

    con = conectar()
    cur = con.cursor()
    cur.execute(
        f"SELECT id, completado FROM viaje_checklist WHERE id = {ph()} AND viaje_id = {ph()}",
        (item_id, viaje_id)
    )
    row = cur.fetchone()
    if not row:
        con.close()
        return jsonify({"error": "not found"}), 404

    nuevo = 0 if row["completado"] else 1
    if nuevo:
        cur.execute(
            "UPDATE viaje_checklist "
            f"SET completado = 1, completado_por = {ph()}, fecha_completado = CURRENT_TIMESTAMP "
            f"WHERE id = {ph()}",
            (session.get("usuario", "admin"), item_id)
        )
    else:
        cur.execute(
            "UPDATE viaje_checklist "
            "SET completado = 0, completado_por = NULL, fecha_completado = NULL "
            f"WHERE id = {ph()}",
            (item_id,)
        )
    con.commit()
    con.close()
    return jsonify({"completado": bool(nuevo)})


@admin_bp.route("/viaje/<int:id>/incidencia/nueva", methods=["POST"])
def nueva_incidencia(id):
    if not requiere_admin():
        return redirect("/login")
    categoria = request.form.get("categoria", "Otro").strip()
    descripcion = request.form.get("descripcion", "").strip()
    if not descripcion:
        return redirect(f"/admin/viaje/{id}#operacion")
    if categoria not in INCIDENCIAS_CATEGORIAS:
        categoria = "Otro"
    con = conectar()
    cur = con.cursor()
    cur.execute(
        f"INSERT INTO incidencias (viaje_id, categoria, descripcion, usuario) VALUES ({ph()}, {ph()}, {ph()}, {ph()})",
        (id, categoria, descripcion, session.get("usuario", "admin"))
    )
    _registrar_historial(cur, id, f"Incidencia registrada: {categoria}", descripcion)
    con.commit()
    con.close()
    return redirect(f"/admin/viaje/{id}#operacion")


@admin_bp.route("/viaje/<int:id>/incidencia/<int:inc_id>/estado", methods=["POST"])
def cambiar_estado_incidencia(id, inc_id):
    if not requiere_admin():
        return jsonify({"error": "no auth"}), 401
    nuevo_estado = request.form.get("estado", "").strip()
    if nuevo_estado not in INCIDENCIAS_ESTADOS:
        return jsonify({"error": "estado inválido"}), 400
    con = conectar()
    cur = con.cursor()
    cur.execute(
        f"UPDATE incidencias SET estado = {ph()} WHERE id = {ph()} AND viaje_id = {ph()}",
        (nuevo_estado, inc_id, id)
    )
    con.commit()
    con.close()
    return jsonify({"estado": nuevo_estado})


@admin_bp.route("/viaje/<int:id>/asignar", methods=["POST"])
def asignar_camionero(id):
    if not requiere_admin():
        return redirect("/login")
    if _viaje_cerrado(id):
        return redirect(f"/admin/viaje/{id}?error=El+viaje+está+cerrado+y+no+admite+cambios")

    camionero_id = request.form["camionero"]

    conexion = conectar()
    cursor = conexion.cursor()

    cursor.execute(f"SELECT nombre FROM camioneros WHERE id = {ph()}", (camionero_id,))
    fila = cursor.fetchone()

    if fila:
        cursor.execute(f"SELECT ruta_id FROM viajes WHERE id = {ph()}", (id,))
        viaje_row = cursor.fetchone()
        ruta_id_directo = viaje_row["ruta_id"] if viaje_row else None
        no_cubiertas = _rutas_no_cubiertas(cursor, camionero_id, id, ruta_id_directo)

        habilitar_rutas = request.form.get("habilitar_rutas") == "1"

        if no_cubiertas and habilitar_rutas:
            for ruta in no_cubiertas:
                asignar_camionero_a_ruta(ruta["ruta_id"], camionero_id)
            etiquetas = ", ".join(f"{r['origen']}–{r['destino']}" for r in no_cubiertas)
            _registrar_historial(
                cursor, id, "Transportista habilitado en rutas",
                f"{fila['nombre']} habilitado para: {etiquetas}"
            )
            no_cubiertas = []

        if no_cubiertas:
            etiquetas = ", ".join(f"{r['origen']}–{r['destino']}" for r in no_cubiertas)
            _registrar_historial(
                cursor, id, "Asignación de transportista rechazada",
                f"{fila['nombre']} no está habilitado para: {etiquetas}"
            )
            conexion.commit()
            conexion.close()
            mensaje = f"El transportista no está habilitado para las rutas: {etiquetas}"
            return redirect(f"/admin/viaje/{id}?error={quote_plus(mensaje)}&camionero_intentado={camionero_id}")

        cursor.execute(f"""
            UPDATE viajes
            SET camionero_id = {ph()}, camionero_nombre = {ph()}, estado = 'Asignado'
            WHERE id = {ph()}
        """, (camionero_id, fila["nombre"], id))

    conexion.commit()

    nombre_camionero = fila['nombre'] if fila else "desconocido"

    conexion.close()

    registrar_auditoria(f"Asignó camionero {nombre_camionero}", "Viajes", "viaje", id, f"Camionero ID: {camionero_id}")

    return redirect(f"/admin/viajes/{id}/gestionar")


@admin_bp.route("/viaje/<int:id>/estado", methods=["POST"])
def cambiar_estado(id):
    if not requiere_admin():
        return redirect("/login")
    if _viaje_cerrado(id):
        return redirect(f"/admin/viaje/{id}?error=El+viaje+está+cerrado+y+no+admite+cambios")

    estado = request.form["estado"]

    conexion = conectar()
    cursor = conexion.cursor()

    cursor.execute(f"""
        SELECT camionero_id, vehiculo_id, precio_final, precio_cliente, precio, estado, cliente_id, fecha_entrega
        FROM viajes WHERE id = {ph()}
    """, (id,))
    viaje = cursor.fetchone()

    if estado == "Asignado":
        if not viaje or not viaje["camionero_id"] or not viaje["vehiculo_id"]:
            conexion.close()
            return redirect(f"/admin/viajes/{id}/gestionar?error=Para+pasar+a+Asignado+debes+asignar+un+camionero+y+un+veh%C3%ADculo")

    if estado == "Entregado":
        if not viaje or not viaje["camionero_id"]:
            conexion.close()
            return redirect(f"/admin/viajes/{id}/gestionar?error=No+puedes+confirmar+la+entrega+sin+un+transportista+asignado")
        conexion.close()
        if tramos_completados(id) is False:
            return redirect(f"/admin/viajes/{id}/gestionar?error=Completa+todos+los+tramos+antes+de+confirmar+la+entrega")
        conexion = conectar()
        cursor = conexion.cursor()

    cursor.execute(f"UPDATE viajes SET estado = {ph()} WHERE id = {ph()}", (estado, id))

    ahora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    entrega_retroactiva = False
    fecha_entrega_previa = None
    if estado == "Asignado":
        cursor.execute(f"UPDATE viajes SET fecha_asignacion = {ph()} WHERE id = {ph()}", (ahora, id))
    elif estado in ["En ruta", "Carga recogida"]:
        cursor.execute(f"UPDATE viajes SET fecha_recogida = {ph()} WHERE id = {ph()}", (ahora, id))
    elif estado == "Entregado":
        fecha_entrega_previa = ((viaje["fecha_entrega"] or "") if viaje else "")[:10]
        hoy = datetime.now().strftime("%Y-%m-%d")
        if not fecha_entrega_previa:
            # No había fecha registrada contra la que comparar: se registra hoy con normalidad.
            cursor.execute(f"UPDATE viajes SET fecha_entrega = {ph()} WHERE id = {ph()}", (ahora, id))
        elif fecha_entrega_previa != hoy:
            # Ya había una fecha de descarga registrada y no coincide con hoy: se respeta esa
            # fecha (no se sobreescribe) y queda marcada como retroactiva en el historial.
            entrega_retroactiva = True

    if estado.lower() in ["entregado", "cancelado"]:
        if viaje and viaje["vehiculo_id"]:
            cursor.execute(
                f"UPDATE vehiculos SET estado = 'Disponible' WHERE id = {ph()}",
                (viaje["vehiculo_id"],)
            )

    email_cliente = None
    if viaje and viaje["cliente_id"]:
        cursor.execute(f"SELECT email FROM clientes WHERE id = {ph()}", (viaje["cliente_id"],))
        cli = cursor.fetchone()
        email_cliente = cli["email"] if cli else None

    estado_anterior = viaje["estado"] if viaje else None

    if entrega_retroactiva:
        _registrar_historial(
            cursor, id, "Entrega confirmada con fecha retroactiva",
            f"Fecha de descarga registrada: {fecha_entrega_previa} · Confirmado el: {ahora}"
        )
    else:
        _registrar_historial(cursor, id, f"Estado cambiado a {estado}",
                             f"Estado anterior: {estado_anterior}" if estado_anterior else "")

    conexion.commit()
    conexion.close()

    if estado in ["Asignado", "En ruta", "Carga recogida", "Entregado", "Cancelado"]:
        import threading
        from flask import current_app
        app_ctx = current_app._get_current_object()
        def _notificar():
            with app_ctx.app_context():
                notificar_cliente_estado(id, estado, email_cliente)
        t = threading.Thread(target=_notificar, daemon=True)
        t.start()

    registrar_auditoria(
        f"Cambió estado a {estado}", "Viajes", "viaje", id,
        f"Estado anterior: {estado_anterior}"
    )

    return redirect(f"/admin/viajes/{id}/gestionar")


@admin_bp.route("/viaje/<int:id>/asignar-todo", methods=["POST"])
def asignar_camionero_vehiculo(id):
    if not requiere_admin():
        return redirect("/login")
    if _viaje_cerrado(id):
        return redirect(f"/admin/viaje/{id}?error=El+viaje+está+cerrado+y+no+admite+cambios")

    camionero_id = request.form.get("camionero", "").strip()

    if not camionero_id:
        return redirect(f"/admin/viajes/{id}/gestionar?error=Selecciona+un+camionero")

    conexion = conectar()
    cursor = conexion.cursor()

    cursor.execute(f"SELECT * FROM viajes WHERE id = {ph()}", (id,))
    viaje = cursor.fetchone()
    if not viaje:
        conexion.close()
        return redirect("/admin/viajes")

    cursor.execute(f"SELECT nombre FROM camioneros WHERE id = {ph()}", (camionero_id,))
    camionero = cursor.fetchone()

    if camionero:
        no_cubiertas = _rutas_no_cubiertas(cursor, camionero_id, id, viaje["ruta_id"])
        habilitar_rutas = request.form.get("habilitar_rutas") == "1"

        if no_cubiertas and habilitar_rutas:
            for ruta in no_cubiertas:
                asignar_camionero_a_ruta(ruta["ruta_id"], camionero_id)
            etiquetas = ", ".join(f"{r['origen']}–{r['destino']}" for r in no_cubiertas)
            _registrar_historial(
                cursor, id, "Transportista habilitado en rutas",
                f"{camionero['nombre']} habilitado para: {etiquetas}"
            )
            no_cubiertas = []

        if no_cubiertas:
            etiquetas = ", ".join(f"{r['origen']}–{r['destino']}" for r in no_cubiertas)
            _registrar_historial(
                cursor, id, "Asignación de transportista rechazada",
                f"{camionero['nombre']} no está habilitado para: {etiquetas}"
            )
            conexion.commit()
            conexion.close()
            mensaje = f"El transportista no está habilitado para las rutas: {etiquetas}"
            return redirect(f"/admin/viaje/{id}?error={quote_plus(mensaje)}&camionero_intentado={camionero_id}")

    cursor.execute(f"""
        SELECT id, COALESCE(matricula, '') AS matricula,
               COALESCE(marca, '') AS marca, COALESCE(modelo, '') AS modelo
        FROM vehiculos
        WHERE camionero_id = {ph()} AND activo = 1
        LIMIT 1
    """, (camionero_id,))
    vehiculo = cursor.fetchone()

    if camionero:
        cursor.execute(f"""
            UPDATE viajes SET camionero_id = {ph()}, camionero_nombre = {ph()} WHERE id = {ph()}
        """, (camionero_id, camionero["nombre"], id))
        cursor.execute(f"UPDATE camioneros SET estado = 'En viaje' WHERE id = {ph()}", (camionero_id,))

    vehiculo_desasignado = False
    if vehiculo:
        cursor.execute(f"""
            UPDATE viajes SET vehiculo_id = {ph()}, vehiculo_placa = {ph()} WHERE id = {ph()}
        """, (vehiculo["id"], vehiculo["matricula"], id))
        cursor.execute(f"UPDATE vehiculos SET estado = 'En viaje' WHERE id = {ph()}", (vehiculo["id"],))
    elif camionero:
        # El nuevo transportista no tiene vehículo activo: no dejar colgado el vehículo anterior
        vehiculo_anterior_id = viaje["vehiculo_id"]
        if vehiculo_anterior_id:
            cursor.execute(f"SELECT estado FROM vehiculos WHERE id = {ph()}", (vehiculo_anterior_id,))
            veh_anterior = cursor.fetchone()
            if veh_anterior and (veh_anterior["estado"] or "") == "En viaje":
                cursor.execute(f"UPDATE vehiculos SET estado = 'Disponible' WHERE id = {ph()}", (vehiculo_anterior_id,))
        cursor.execute(f"UPDATE viajes SET vehiculo_id = NULL, vehiculo_placa = NULL WHERE id = {ph()}", (id,))
        vehiculo_desasignado = True

    if camionero:
        cursor.execute(f"UPDATE viajes SET estado = 'Asignado' WHERE id = {ph()}", (id,))

    nombre_cam = camionero["nombre"] if camionero else "desconocido"
    nombre_veh = f"{vehiculo['marca']} {vehiculo['modelo']} ({vehiculo['matricula']})" if vehiculo else "desconocido"

    _registrar_historial(cursor, id, f"Camionero asignado: {nombre_cam}", f"Vehículo: {nombre_veh}")
    if vehiculo_desasignado:
        _registrar_historial(
            cursor, id, "Vehículo desasignado",
            f"El viaje quedó sin vehículo asignado tras reasignar a {nombre_cam} (sin vehículo activo)"
        )

    conexion.commit()
    conexion.close()

    registrar_auditoria(
        f"Asignó camionero {nombre_cam} y vehículo {nombre_veh}",
        "Viajes", "viaje", id,
        f"Camionero y vehículo asignados juntos"
    )

    return redirect(f"/admin/viajes/{id}/gestionar")


@admin_bp.route("/viaje/<int:id>/pdf")
def descargar_pdf_orden_carga(id):
    if not requiere_admin():
        return redirect("/login")

    conexion = conectar()
    cursor = conexion.cursor()

    cursor.execute(f"""
        SELECT
            v.*,
            c.telefono  AS camionero_telefono,
            c.licencia  AS camionero_licencia,
            veh.matricula AS vehiculo_matricula,
            veh.marca     AS vehiculo_marca,
            veh.modelo    AS vehiculo_modelo,
            cl.empresa    AS cliente_empresa,
            cl.telefono   AS cliente_telefono,
            cl.email      AS cliente_email
        FROM viajes v
        LEFT JOIN camioneros c   ON v.camionero_id = c.id
        LEFT JOIN vehiculos  veh ON v.vehiculo_id  = veh.id
        LEFT JOIN clientes   cl  ON v.cliente_id   = cl.id
        WHERE v.id = {ph()}
    """, (id,))

    fila = cursor.fetchone()
    conexion.close()

    if not fila:
        return redirect("/admin/viajes")

    viaje = dict(fila)

    faltantes_oc = []
    if not viaje.get("cliente"):
        faltantes_oc.append("cliente")
    if not viaje.get("camionero_nombre") and not viaje.get("camionero_id"):
        faltantes_oc.append("camionero")
    if not viaje.get("vehiculo_id"):
        faltantes_oc.append("vehículo")
    if not viaje.get("origen"):
        faltantes_oc.append("origen")
    if not viaje.get("destino"):
        faltantes_oc.append("destino")
    if faltantes_oc:
        msg = quote_plus("Faltan datos para Orden de Carga: " + ", ".join(faltantes_oc))
        return redirect(f"/admin/viajes/{id}/gestionar?error={msg}")

    try:
        pdf_bytes = generar_pdf_orden_carga(viaje)
    except ValueError as e:
        msg = quote_plus(str(e))
        return redirect(f"/admin/viajes/{id}/gestionar?error={msg}")
    except Exception as e:
        from flask import current_app
        current_app.logger.error(f"Error PDF orden carga viaje {id}: {e}")
        msg = quote_plus("Error generando PDF de Orden de Carga")
        return redirect(f"/admin/viajes/{id}/gestionar?error={msg}")

    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype="application/pdf",
        as_attachment=True,
        download_name=f"orden-carga-{id:04d}.pdf",
    )


@admin_bp.route("/viaje/<int:id>/carta-porte")
def descargar_carta_porte(id):
    if not requiere_admin():
        return redirect("/login")
    try:
        from services.pdf_service import generar_pdf_carta_porte
        pdf_bytes = generar_pdf_carta_porte(id)
        registrar_auditoria("Descargó Carta de Porte", "Viajes", "viaje", id)
        return send_file(
            io.BytesIO(pdf_bytes),
            mimetype="application/pdf",
            as_attachment=True,
            download_name=f"carta-porte-{id:04d}.pdf",
        )
    except ValueError as e:
        msg = quote_plus(str(e))
        return redirect(f"/admin/viajes/{id}/gestionar?error={msg}")
    except Exception as e:
        from flask import current_app
        current_app.logger.error(f"Error PDF carta porte viaje {id}: {e}")
        msg = quote_plus("Error generando Carta de Porte")
        return redirect(f"/admin/viajes/{id}/gestionar?error={msg}")


@admin_bp.route("/viaje/<int:id>/liquidacion")
def descargar_liquidacion(id):
    if session.get("rol") != "admin":
        return redirect("/login")
    try:
        from services.pdf_service import generar_pdf_liquidacion_camionero
        pdf_bytes = generar_pdf_liquidacion_camionero(id)
        return send_file(
            io.BytesIO(pdf_bytes),
            mimetype="application/pdf",
            as_attachment=True,
            download_name=f"liquidacion-{id:04d}.pdf",
        )
    except ValueError as e:
        msg = quote_plus(str(e))
        return redirect(f"/admin/viajes/{id}/gestionar?error={msg}")
    except Exception as e:
        from flask import current_app
        current_app.logger.error(f"Error PDF liquidacion viaje {id}: {e}")
        msg = quote_plus("Error generando Liquidación")
        return redirect(f"/admin/viajes/{id}/gestionar?error={msg}")


@admin_bp.route("/viaje/<int:id>/factura")
def descargar_factura_cliente(id):
    if not requiere_admin():
        return redirect("/login")

    try:
        pdf_bytes = generar_factura_cliente(id)
    except ValueError as e:
        msg = quote_plus(str(e))
        return redirect(f"/admin/viajes/{id}/gestionar?error={msg}")
    except Exception as e:
        from flask import current_app
        current_app.logger.error(f"Error PDF factura viaje {id}: {e}")
        msg = quote_plus("Error generando Factura")
        return redirect(f"/admin/viajes/{id}/gestionar?error={msg}")

    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype="application/pdf",
        as_attachment=True,
        download_name=f"factura-{id:04d}.pdf",
    )


@admin_bp.route("/viaje/<int:id>/guardar-combustible", methods=["POST"])
def guardar_combustible(id):
    if not requiere_admin():
        return redirect("/login")
    if _viaje_cerrado(id):
        return redirect(f"/admin/viaje/{id}?error=El+viaje+está+cerrado+y+no+admite+cambios")
    val = request.form.get("combustible", "").strip()
    try:
        combustible = float(val)
        if combustible < 0:
            raise ValueError
    except (ValueError, TypeError):
        return redirect(f"/admin/viaje/{id}?error=Valor+de+combustible+inv%C3%A1lido")
    con = conectar()
    cur = con.cursor()
    cur.execute(f"UPDATE viajes SET combustible = {ph()} WHERE id = {ph()}", (combustible, id))
    con.commit()
    con.close()
    registrar_auditoria(f"Guardó combustible ${combustible:.2f}", "Viajes", "viaje", id)
    return redirect(f"/admin/viaje/{id}")


@admin_bp.route("/viaje/<int:id>/guardar-fechas", methods=["POST"])
def guardar_fechas(id):
    if not requiere_admin():
        return redirect("/login")
    if _viaje_cerrado(id):
        return redirect(f"/admin/viaje/{id}?error=El+viaje+está+cerrado+y+no+admite+cambios")
    fecha_recogida = request.form.get("fecha_recogida", "").strip() or None
    fecha_entrega = request.form.get("fecha_entrega", "").strip() or None
    con = conectar()
    cur = con.cursor()
    if fecha_recogida:
        cur.execute(f"UPDATE viajes SET fecha_recogida = {ph()} WHERE id = {ph()}", (fecha_recogida, id))
    if fecha_entrega:
        cur.execute(f"UPDATE viajes SET fecha_entrega = {ph()} WHERE id = {ph()}", (fecha_entrega, id))
    con.commit()
    con.close()
    return redirect(f"/admin/viaje/{id}")


@admin_bp.route("/incidencias")
def lista_incidencias():
    if not requiere_admin():
        return redirect("/login")
    filtro_estado = request.args.get("estado", "").strip()
    filtro_cat = request.args.get("categoria", "").strip()
    con = conectar()
    cur = con.cursor()
    conds = []
    params = []
    if filtro_estado:
        conds.append(f"i.estado = {ph()}")
        params.append(filtro_estado)
    if filtro_cat:
        conds.append(f"i.categoria = {ph()}")
        params.append(filtro_cat)
    where = ("WHERE " + " AND ".join(conds)) if conds else ""
    cur.execute(f"""
        SELECT i.id, i.viaje_id, i.categoria, i.descripcion,
               i.usuario, i.fecha_hora, i.estado,
               v.origen, v.destino, v.estado AS viaje_estado, v.cliente
        FROM incidencias i
        LEFT JOIN viajes v ON i.viaje_id = v.id
        {where}
        ORDER BY i.fecha_hora DESC
        LIMIT 500
    """, params)
    incidencias = cur.fetchall()
    con.close()
    return render_template("admin/incidencias.html",
                           incidencias=incidencias,
                           filtro_estado=filtro_estado,
                           filtro_cat=filtro_cat,
                           categorias=INCIDENCIAS_CATEGORIAS,
                           estados=INCIDENCIAS_ESTADOS)


@admin_bp.route("/viaje/<int:id>/eliminar", methods=["POST"])
def eliminar_viaje_admin(id):
    if session.get("rol") != "admin":
        return redirect("/admin/viajes")
    conexion = conectar()
    cursor = conexion.cursor()
    cursor.execute(f"""
        UPDATE viajes SET deleted_at = CURRENT_TIMESTAMP, deleted_by = {ph()}
        WHERE id = {ph()}
    """, (session.get("usuario"), id))
    conexion.commit()
    conexion.close()
    registrar_auditoria("eliminó (papelera)", "viajes", "viaje", id)
    return redirect("/admin/viajes")


@admin_bp.route("/viaje/<int:id>/prioridad", methods=["POST"])
def actualizar_prioridad_viaje(id):
    if not requiere_admin():
        return redirect("/login")
    if _viaje_cerrado(id):
        return redirect(f"/admin/viaje/{id}?error=El+viaje+está+cerrado+y+no+admite+cambios")
    prioridad = request.form.get("prioridad", "Normal").strip()
    if prioridad not in ["Normal", "Alta", "Urgente"]:
        prioridad = "Normal"
    conexion = conectar()
    cursor = conexion.cursor()
    cursor.execute(f"UPDATE viajes SET prioridad = {ph()} WHERE id = {ph()}", (prioridad, id))
    conexion.commit()
    conexion.close()
    return redirect(f"/admin/viaje/{id}#operacion")


@admin_bp.route("/viaje/<int:id>/nota", methods=["POST"])
def agregar_nota_viaje(id):
    if not requiere_admin():
        return redirect("/login")
    texto = request.form.get("texto", "").strip()
    if texto:
        conexion = conectar()
        cursor = conexion.cursor()
        cursor.execute(
            f"INSERT INTO notas_viaje (viaje_id, usuario, texto) VALUES ({ph()}, {ph()}, {ph()})",
            (id, session.get("usuario", "admin"), texto)
        )
        conexion.commit()
        conexion.close()
    return redirect(f"/admin/viaje/{id}")


@admin_bp.route("/cotizacion/<int:id>/convertir")
@admin_bp.route("/cotizaciones/<int:id>/convertir")
def convertir_cotizacion(id):
    if not requiere_admin():
        return redirect("/login")

    viaje_id = convertir_cotizacion_en_viaje(id)

    if not viaje_id:
        return redirect("/comercial/cotizaciones")

    return redirect(f"/admin/viajes/{viaje_id}/gestionar")


@admin_bp.route("/viaje/<int:id>/pago-camionero", methods=["POST"])
def pago_camionero(id):
    if not requiere_admin():
        return redirect("/login")
    if _viaje_cerrado(id):
        return redirect(f"/admin/viaje/{id}?error=El+viaje+está+cerrado+y+no+admite+cambios")

    accion     = request.form.get("accion", "").strip()
    tipo_pago  = request.form.get("tipo_pago", "").strip() or None
    observacion = request.form.get("observacion", "").strip() or None
    monto_str  = request.form.get("monto_parcial", "").strip()

    try:
        monto = float(monto_str) if monto_str else None
    except ValueError:
        monto = None

    con = conectar()
    cur = con.cursor()

    auditoria_msg = None

    if accion == "pagado" and session.get("rol") == "admin":
        cur.execute(
            f"UPDATE viajes SET estado_pago_camionero='Pagado', tipo_pago_camionero={ph()}, "
            f"observacion_pago={ph()}, fecha_pago_camionero=CURRENT_TIMESTAMP WHERE id={ph()}",
            (tipo_pago, observacion, id)
        )
        auditoria_msg = "Marcó pago camionero como Pagado"

    elif accion == "parcial" and session.get("rol") == "admin":
        cur.execute(
            f"UPDATE viajes SET estado_pago_camionero='Parcial', tipo_pago_camionero={ph()}, "
            f"observacion_pago={ph()}, monto_pagado={ph()} WHERE id={ph()}",
            (tipo_pago, observacion, monto, id)
        )
        auditoria_msg = "Marcó pago camionero como Parcial"

    elif accion == "revertir" and session.get("rol") == "admin":
        cur.execute(
            f"UPDATE viajes SET estado_pago_camionero='Pendiente', fecha_pago_camionero=NULL WHERE id={ph()}",
            (id,)
        )

    con.commit()
    con.close()

    if auditoria_msg:
        registrar_auditoria(auditoria_msg, "Viajes", "viaje", id)

    return redirect("/admin/pagos-pendientes")


@admin_bp.route("/pagos-pendientes")
def pagos_pendientes():
    if session.get("rol") != "admin":
        return redirect("/login")

    conexion = conectar()
    cursor = conexion.cursor()
    cursor.execute(f"""
        SELECT v.id, v.origen, v.destino, v.camionero_nombre,
               v.estado_pago_camionero, v.monto_pagado, v.observacion_pago, v.tipo_pago_camionero
        FROM viajes v
        WHERE LOWER(v.estado) IN ('entregado', 'cerrado')
          AND (v.estado_pago_camionero IS NULL OR v.estado_pago_camionero != 'Pagado')
          AND v.deleted_at IS NULL
        ORDER BY v.id DESC
    """)
    filas = cursor.fetchall()
    conexion.close()

    pendientes = []
    for v in filas:
        liquidacion = calcular_liquidacion(v["id"])
        monto_calculado = liquidacion["pago_camionero"] if liquidacion else 0
        pendientes.append({
            "id": v["id"],
            "origen": v["origen"],
            "destino": v["destino"],
            "camionero_nombre": v["camionero_nombre"],
            "monto_calculado": monto_calculado,
            "estado_pago_camionero": v["estado_pago_camionero"],
            "monto_pagado": v["monto_pagado"],
            "observacion_pago": v["observacion_pago"],
            "tipo_pago_camionero": v["tipo_pago_camionero"],
        })

    return render_template("admin/pagos_pendientes.html", pendientes=pendientes)


@admin_bp.route("/viaje/<int:id>/marcar-cobrado", methods=["POST"])
def marcar_cobrado(id):
    if session.get("rol") != "admin":
        return redirect("/login")
    if _viaje_cerrado(id):
        return redirect(f"/admin/viaje/{id}?error=El+viaje+está+cerrado+y+no+admite+cambios")

    forma_cobro = request.form.get("forma_cobro", "").strip()
    codigo_transaccion = request.form.get("codigo_transaccion", "").strip() or None
    comentario_cobro = request.form.get("comentario_cobro", "").strip() or None
    monto_str = request.form.get("monto_cobrado", "").strip()
    try:
        monto_cobrado = float(monto_str) if monto_str else None
    except ValueError:
        monto_cobrado = None

    con = conectar()
    cur = con.cursor()

    cur.execute(f"SELECT camionero_id FROM viajes WHERE id = {ph()}", (id,))
    viaje = cur.fetchone()
    if not viaje or not viaje["camionero_id"]:
        con.close()
        return redirect(f"/admin/viajes/{id}/gestionar?error=No+puedes+registrar+el+cobro+sin+un+transportista+asignado")

    cur.execute(
        f"UPDATE viajes SET forma_cobro={ph()}, codigo_transaccion={ph()}, "
        f"comentario_cobro={ph()}, fecha_cobro=CURRENT_TIMESTAMP, monto_cobrado={ph()} "
        f"WHERE id={ph()}",
        (forma_cobro, codigo_transaccion, comentario_cobro, monto_cobrado, id)
    )

    detalle = f"Forma: {forma_cobro}"
    if monto_cobrado:
        detalle += f" · Monto: ${monto_cobrado:.2f}"
    if codigo_transaccion:
        detalle += f" · Código: {codigo_transaccion}"
    _registrar_historial(cur, id, "Cobro registrado", detalle)

    con.commit()
    con.close()

    registrar_auditoria("Registró cobro del viaje", "Viajes", "viaje", id, detalle)

    referer = request.form.get("_referer", "")
    if referer and "reportes" in referer:
        return redirect(referer)
    return redirect(f"/admin/viajes/{id}/gestionar")


@admin_bp.route("/viaje/<int:id>/finalizar", methods=["POST"])
def finalizar_viaje(id):
    if session.get("rol") != "admin":
        return redirect("/login")
    if _viaje_cerrado(id):
        return redirect(f"/admin/viaje/{id}")

    con = conectar()
    cur = con.cursor()
    cur.execute(f"SELECT fecha_cobro FROM viajes WHERE id = {ph()}", (id,))
    viaje = cur.fetchone()
    if not viaje or not viaje["fecha_cobro"]:
        con.close()
        return redirect(f"/admin/viaje/{id}?error=El+cobro+debe+estar+registrado+antes+de+finalizar+el+viaje")

    cur.execute(f"UPDATE viajes SET estado = 'Cerrado' WHERE id = {ph()}", (id,))
    _registrar_historial(cur, id, "Viaje finalizado", "Estado cambiado a Cerrado")
    con.commit()
    con.close()

    registrar_auditoria("Finalizó el viaje", "Viajes", "viaje", id)

    return redirect(f"/admin/viaje/{id}")


@admin_bp.route("/viaje/<int:id>/reabrir", methods=["POST"])
def reabrir_viaje(id):
    if session.get("rol") != "admin":
        return redirect("/login")
    if not _viaje_cerrado(id):
        return redirect(f"/admin/viaje/{id}")

    con = conectar()
    cur = con.cursor()
    cur.execute(f"UPDATE viajes SET estado = 'Entregado', reabierto_en = CURRENT_TIMESTAMP WHERE id = {ph()}", (id,))
    _registrar_historial(cur, id, "Viaje reabierto", "Estado cambiado de Cerrado a Entregado")
    con.commit()
    con.close()

    registrar_auditoria("Reabrió el viaje", "Viajes", "viaje", id)

    return redirect(f"/admin/viaje/{id}")


@admin_bp.route("/viaje/<int:id>/corregir-cobro", methods=["POST"])
def corregir_cobro(id):
    if session.get("rol") != "admin":
        return redirect("/login")

    con = conectar()
    cur = con.cursor()
    cur.execute(f"""
        SELECT reabierto_en, estado, forma_cobro, monto_cobrado, codigo_transaccion,
               verificado_financiero
        FROM viajes WHERE id = {ph()}
    """, (id,))
    viaje = cur.fetchone()
    if not viaje or not viaje["reabierto_en"] or (viaje["estado"] or "").lower() == "cerrado":
        con.close()
        return redirect(f"/admin/viaje/{id}?error=La+corrección+de+cobro+solo+aplica+a+viajes+reabiertos")

    forma_cobro_nueva = request.form.get("forma_cobro", "").strip()
    codigo_nuevo = request.form.get("codigo_transaccion", "").strip() or None
    monto_str = request.form.get("monto_cobrado", "").strip()
    try:
        monto_nuevo = float(monto_str) if monto_str else None
    except ValueError:
        monto_nuevo = None

    forma_anterior = viaje["forma_cobro"] or ""
    monto_anterior = float(viaje["monto_cobrado"] or 0)
    codigo_anterior = viaje["codigo_transaccion"] or None

    cambios = []
    if forma_cobro_nueva and forma_cobro_nueva != forma_anterior:
        cambios.append(f"forma de pago: '{forma_anterior or '—'}' → '{forma_cobro_nueva}'")
    if monto_nuevo is not None and monto_nuevo != monto_anterior:
        cambios.append(f"monto: ${monto_anterior:.2f} → ${monto_nuevo:.2f}")
    if codigo_nuevo != codigo_anterior:
        cambios.append(f"código transacción: '{codigo_anterior or '—'}' → '{codigo_nuevo or '—'}'")

    cur.execute(f"""
        UPDATE viajes SET forma_cobro = {ph()}, monto_cobrado = {ph()}, codigo_transaccion = {ph()}
        WHERE id = {ph()}
    """, (
        forma_cobro_nueva or viaje["forma_cobro"],
        monto_nuevo if monto_nuevo is not None else viaje["monto_cobrado"],
        codigo_nuevo,
        id,
    ))

    if cambios and viaje["verificado_financiero"]:
        cur.execute(
            f"UPDATE viajes SET verificado_financiero=0, verificado_por=NULL, fecha_verificacion=NULL WHERE id={ph()}",
            (id,)
        )
        cambios.append("verificación financiera revertida a pendiente por la corrección")

    if cambios:
        _registrar_historial(cur, id, "Cobro corregido", "; ".join(cambios))

    con.commit()
    con.close()

    if cambios:
        registrar_auditoria("Corrigió el cobro del viaje", "Viajes", "viaje", id, "; ".join(cambios))

    return redirect(f"/admin/viaje/{id}")


@admin_bp.route("/viaje/<int:id>/verificar", methods=["POST"])
def verificar_viaje(id):
    if session.get("rol") != "admin":
        return redirect("/login")
    accion = request.form.get("accion", "verificar")
    con = conectar()
    cur = con.cursor()
    if accion == "revertir":
        cur.execute(
            f"UPDATE viajes SET verificado_financiero=0, verificado_por=NULL, fecha_verificacion=NULL WHERE id={ph()}",
            (id,)
        )
        _registrar_historial(cur, id, "Verificación revertida", f"Por: {session.get('usuario')}")
    else:
        ahora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cur.execute(
            f"UPDATE viajes SET verificado_financiero=1, verificado_por={ph()}, fecha_verificacion={ph()} WHERE id={ph()}",
            (session.get("usuario"), ahora, id)
        )
        _registrar_historial(cur, id, "Verificado financiero", f"Por: {session.get('usuario')}")
    con.commit()
    con.close()

    referer = request.form.get("_referer", "/admin/reportes")
    return redirect(referer)


@admin_bp.route("/camioneros/<int:id>/economico")
def camionero_economico(id):
    if not requiere_admin():
        return redirect("/login")

    con = conectar()
    cur = con.cursor()

    cur.execute(
        f"SELECT id, nombre, telefono, licencia, estado FROM camioneros WHERE id = {ph()}", (id,)
    )
    camionero = cur.fetchone()
    if not camionero:
        con.close()
        return redirect("/admin/camioneros")

    cur.execute(f"""
        SELECT id, fecha_creacion, origen, destino, estado,
               tipo_pago_camionero, estado_pago_camionero, monto_pagado,
               observacion_pago, fecha_pago_camionero
        FROM viajes
        WHERE camionero_id = {ph()} AND LOWER(estado) != 'cancelado'
        ORDER BY id DESC
    """, (id,))
    viajes_camionero = cur.fetchall()
    con.close()

    total_generado = 0.0
    total_pagado = 0.0
    pendientes = []
    pagados = []

    for v in viajes_camionero:
        liquidacion = calcular_liquidacion(v["id"])
        monto_calculado = liquidacion["pago_camionero"] if liquidacion else 0
        total_generado += monto_calculado

        estado_pago = v["estado_pago_camionero"]
        fila = {
            "id": v["id"],
            "fecha_creacion": v["fecha_creacion"],
            "origen": v["origen"],
            "destino": v["destino"],
            "estado": v["estado"],
            "monto_calculado": monto_calculado,
            "tipo_pago_camionero": v["tipo_pago_camionero"],
            "estado_pago_camionero": estado_pago,
            "monto_pagado": v["monto_pagado"],
            "observacion_pago": v["observacion_pago"],
            "fecha_pago_camionero": v["fecha_pago_camionero"],
        }

        if estado_pago == "Pagado":
            total_pagado += monto_calculado
            pagados.append(fila)
        else:
            if estado_pago == "Parcial":
                total_pagado += float(v["monto_pagado"] or 0)
            pendientes.append(fila)

    pagados = pagados[:10]
    pendiente = total_generado - total_pagado

    return render_template(
        "admin/camionero_economico.html",
        camionero=camionero,
        total_generado=total_generado,
        total_pagado=total_pagado,
        pendiente=pendiente,
        pendientes=pendientes,
        pagados=pagados,
        es_admin=(session.get("rol") == "admin"),
    )


@admin_bp.route("/catalogos/tipo-transporte", methods=["GET", "POST"])
def catalogo_tipo_transporte_admin():
    if not requiere_admin():
        return redirect("/login")
    if session.get("rol") != "admin":
        return redirect("/admin")

    con = conectar()
    cur = con.cursor()
    error = None

    if request.method == "POST":
        accion = request.form.get("accion", "")
        if accion == "nuevo":
            nombre = request.form.get("nombre", "").strip()
            if not nombre:
                error = "El nombre es obligatorio."
            else:
                try:
                    cur.execute(
                        f"INSERT INTO catalogo_tipo_transporte (nombre) VALUES ({ph()})", (nombre,)
                    )
                    con.commit()
                except Exception:
                    con.rollback()
                    error = f"'{nombre}' ya existe en el catálogo."
        elif accion == "toggle":
            tipo_id = request.form.get("tipo_id", "")
            if tipo_id:
                cur.execute(
                    f"SELECT activo FROM catalogo_tipo_transporte WHERE id = {ph()}", (tipo_id,)
                )
                row = cur.fetchone()
                if row:
                    cur.execute(
                        f"UPDATE catalogo_tipo_transporte SET activo = {ph()} WHERE id = {ph()}",
                        (0 if row["activo"] else 1, tipo_id)
                    )
                    con.commit()
        if not error:
            con.close()
            return redirect("/admin/catalogos/tipo-transporte")

    cur.execute("SELECT id, nombre, activo FROM catalogo_tipo_transporte ORDER BY nombre")
    tipos = cur.fetchall()
    con.close()

    return render_template(
        "admin/catalogo_tipo_transporte.html",
        tipos=tipos,
        error=error,
    )


# ── Camioneros CRUD ──────────────────────────────────────────────────────────

@admin_bp.route("/camioneros", methods=["GET", "POST"])
def admin_camioneros():
    if not requiere_admin():
        return redirect("/login")

    conexion = conectar()
    cursor = conexion.cursor()

    error = None

    if request.method == "POST":
        nombre             = request.form["nombre"].strip()
        telefono           = request.form.get("telefono", "").strip()
        licencia           = request.form.get("licencia", "").strip()
        carnet_identidad   = request.form.get("carnet_identidad", "").strip()
        licencia_operativa = request.form.get("licencia_operativa", "").strip()
        empresa            = request.form.get("empresa", "").strip()
        estado             = request.form.get("estado", "Disponible").strip()

        matricula      = request.form.get("matricula", "").strip()
        marca          = request.form.get("marca", "").strip()
        modelo         = request.form.get("modelo", "").strip()
        tipo           = request.form.get("tipo", "").strip()
        capacidad      = request.form.get("capacidad", "").strip()
        chapa_remolque = request.form.get("chapa_remolque", "").strip()

        if not matricula:
            error = "La matrícula del vehículo es obligatoria."

        if not error:
            cursor.execute(f"SELECT id FROM vehiculos WHERE matricula = {ph()}", (matricula,))
            if cursor.fetchone():
                error = f"La matrícula '{matricula}' ya está registrada."

        if not error:
            cursor.execute(f"""
                INSERT INTO camioneros (nombre, telefono, licencia, carnet_identidad,
                    licencia_operativa, empresa, estado, activo)
                VALUES ({ph()}, {ph()}, {ph()}, {ph()}, {ph()}, {ph()}, {ph()}, 1)
            """, (nombre, telefono, licencia, carnet_identidad, licencia_operativa, empresa, estado))
            nuevo_id = cursor.lastrowid

            if matricula:
                cursor.execute(f"""
                    INSERT INTO vehiculos
                        (camionero_id, matricula, marca, modelo, tipo, capacidad,
                         chapa_remolque, estado, activo)
                    VALUES ({ph()}, {ph()}, {ph()}, {ph()}, {ph()}, {ph()}, {ph()}, 'Disponible', 1)
                """, (nuevo_id, matricula, marca, modelo, tipo, capacidad, chapa_remolque))

            conexion.commit()
            conexion.close()
            return redirect("/admin/camioneros?ok=1")

    buscar           = request.args.get("buscar", "").strip()
    filtro_estado    = request.args.get("estado", "").strip()
    filtro_tipo      = request.args.get("tipo_transporte", "").strip()
    pagina           = max(1, int(request.args.get("pagina", 1) or 1))
    por_pagina       = 20

    condiciones = ["(c.activo = 1 OR c.activo IS NULL)", "c.deleted_at IS NULL"]
    params = []
    if buscar:
        condiciones.append(f"(c.nombre LIKE {ph()} OR c.telefono LIKE {ph()} OR c.licencia LIKE {ph()})")
        like = f"%{buscar}%"
        params.extend([like, like, like])
    if filtro_estado:
        condiciones.append(f"LOWER(c.estado) = {ph()}")
        params.append(filtro_estado.lower())
    if filtro_tipo:
        condiciones.append(f"LOWER(COALESCE(v.tipo, '')) = {ph()}")
        params.append(filtro_tipo.lower())

    where = "WHERE " + " AND ".join(condiciones)

    cursor.execute(f"""
        SELECT COUNT(DISTINCT c.id) AS total
        FROM camioneros c
        LEFT JOIN vehiculos v ON v.camionero_id = c.id AND v.activo = 1
        {where}
    """, params)
    total = cursor.fetchone()["total"]
    total_paginas = max(1, (total + por_pagina - 1) // por_pagina)
    pagina = min(pagina, total_paginas)
    offset = (pagina - 1) * por_pagina

    cursor.execute(f"""
        SELECT c.id, c.nombre, c.telefono, c.licencia, c.estado,
               COALESCE(v.matricula, '—') AS vehiculo_matricula,
               COALESCE(v.marca, '') AS vehiculo_marca,
               COALESCE(v.modelo, '') AS vehiculo_modelo,
               COALESCE(v.tipo, '') AS vehiculo_tipo,
               v.id AS vehiculo_id
        FROM camioneros c
        LEFT JOIN vehiculos v ON v.camionero_id = c.id AND v.activo = 1
        {where}
        ORDER BY c.id DESC
        LIMIT ? OFFSET ?
    """, params + [por_pagina, offset])
    lista = cursor.fetchall()

    cursor.execute(
        "SELECT id, nombre FROM tipos_vehiculo WHERE activo = 1 ORDER BY nombre"
    )
    tipos_vehiculo = cursor.fetchall()

    conexion.close()

    rutas_por_camionero = {c["id"]: get_rutas_por_camionero(c["id"]) for c in lista}

    return render_template(
        "admin/camioneros.html",
        lista=lista,
        estados=CAMIONERO_ESTADOS,
        rutas_por_camionero=rutas_por_camionero,
        buscar=buscar,
        filtro_estado=filtro_estado,
        filtro_tipo=filtro_tipo,
        tipos_vehiculo=tipos_vehiculo,
        pagina_actual=pagina,
        total_paginas=total_paginas,
        total=total,
        error=error,
        form_data=request.form if error else None,
    )


@admin_bp.route("/camioneros/<int:id>/editar", methods=["GET", "POST"])
def editar_camionero(id):
    if not requiere_admin():
        return redirect("/login")

    conexion = conectar()
    cursor = conexion.cursor()

    error = None

    if request.method == "POST":
        nombre             = request.form["nombre"].strip()
        telefono           = request.form.get("telefono", "").strip()
        licencia           = request.form.get("licencia", "").strip()
        carnet_identidad   = request.form.get("carnet_identidad", "").strip()
        licencia_operativa = request.form.get("licencia_operativa", "").strip()
        empresa            = request.form.get("empresa", "").strip()
        estado             = request.form.get("estado", "Disponible").strip()

        matricula      = request.form.get("matricula", "").strip()
        marca          = request.form.get("marca", "").strip()
        modelo         = request.form.get("modelo", "").strip()
        tipo           = request.form.get("tipo", "").strip()
        capacidad      = request.form.get("capacidad", "").strip()
        chapa_remolque = request.form.get("chapa_remolque", "").strip()

        cursor.execute(f"""
            UPDATE camioneros SET nombre = {ph()}, telefono = {ph()}, licencia = {ph()},
                carnet_identidad = {ph()}, licencia_operativa = {ph()}, empresa = {ph()},
                estado = {ph()}
            WHERE id = {ph()}
        """, (nombre, telefono, licencia, carnet_identidad, licencia_operativa, empresa, estado, id))

        cursor.execute(
            f"SELECT id FROM vehiculos WHERE camionero_id = {ph()} AND activo = 1", (id,)
        )
        veh_row = cursor.fetchone()

        if veh_row:
            cursor.execute(f"""
                UPDATE vehiculos
                SET matricula = {ph()}, marca = {ph()}, modelo = {ph()}, tipo = {ph()},
                    capacidad = {ph()}, chapa_remolque = {ph()}
                WHERE id = {ph()}
            """, (matricula, marca, modelo, tipo, capacidad, chapa_remolque, veh_row["id"]))
        elif matricula:
            cursor.execute(
                f"SELECT id FROM vehiculos WHERE matricula = {ph()}", (matricula,)
            )
            if cursor.fetchone():
                error = f"La matrícula '{matricula}' ya está registrada."
            else:
                cursor.execute(f"""
                    INSERT INTO vehiculos
                        (camionero_id, matricula, marca, modelo, tipo, capacidad,
                         chapa_remolque, estado, activo)
                    VALUES ({ph()}, {ph()}, {ph()}, {ph()}, {ph()}, {ph()}, {ph()}, 'Disponible', 1)
                """, (id, matricula, marca, modelo, tipo, capacidad, chapa_remolque))

        if not error:
            conexion.commit()
            conexion.close()
            return redirect("/admin/camioneros")

    cursor.execute(
        f"""SELECT id, nombre, telefono, licencia, carnet_identidad,
                   licencia_operativa, empresa, estado
            FROM camioneros WHERE id = {ph()}""",
        (id,)
    )
    camionero = cursor.fetchone()

    if not camionero:
        conexion.close()
        return redirect("/admin/camioneros")

    cursor.execute(
        f"""SELECT id, matricula, marca, modelo, tipo, capacidad, chapa_remolque
           FROM vehiculos WHERE camionero_id = {ph()} AND activo = 1""",
        (id,)
    )
    vehiculo = cursor.fetchone()

    cursor.execute(
        "SELECT id, nombre FROM tipos_vehiculo WHERE activo = 1 ORDER BY nombre"
    )
    tipos_vehiculo = cursor.fetchall()
    conexion.close()

    return render_template(
        "admin/editar_camionero.html",
        camionero=camionero,
        vehiculo=vehiculo,
        estados=CAMIONERO_ESTADOS,
        tipos_vehiculo=tipos_vehiculo,
        error=error,
    )


@admin_bp.route("/camioneros/<int:id>/eliminar", methods=["POST"])
def eliminar_camionero(id):
    if not requiere_admin():
        return redirect("/login")

    conexion = conectar()
    cursor = conexion.cursor()

    if session.get("rol") == "admin":
        # Soft delete directo
        cursor.execute(f"SELECT nombre FROM camioneros WHERE id = {ph()}", (id,))
        row = cursor.fetchone()
        cursor.execute(f"""
            UPDATE camioneros SET deleted_at = CURRENT_TIMESTAMP, deleted_by = {ph()}
            WHERE id = {ph()}
        """, (session.get("usuario"), id))
        conexion.commit()
        conexion.close()
        registrar_auditoria("eliminó (papelera)", "camioneros", "camionero", id)
        return redirect("/admin/camioneros")
    else:
        # Operario: crea solicitud de aprobación
        cursor.execute(f"SELECT nombre FROM camioneros WHERE id = {ph()}", (id,))
        row = cursor.fetchone()
        nombre_entidad = row["nombre"] if row else f"#{id}"
        cursor.execute(f"""
            INSERT INTO solicitudes_eliminacion
                (entidad, entidad_id, nombre_entidad, solicitado_por)
            VALUES ({ph()}, {ph()}, {ph()}, {ph()})
        """, ("camionero", id, nombre_entidad, session.get("usuario")))
        conexion.commit()
        conexion.close()
        return redirect("/admin/camioneros?access_error=Solicitud+de+eliminaci%C3%B3n+enviada+al+administrador")


# ── Clientes CRUD ────────────────────────────────────────────────────────────

CATEGORIAS_CLIENTE = ["Normal", "VIP", "Estratégico", "Humanitario"]


def _validar_y_crear_cliente(cursor, form):
    """Valida y crea un cliente a partir de un form (request.form).
    Devuelve (cliente_id, error) — error es None si se creó correctamente."""
    nombre = form["nombre"].strip()
    empresa = form.get("empresa", "").strip()
    contacto = form.get("contacto", "").strip()
    telefono = form.get("telefono", "").strip()
    email = form.get("email", "").strip()
    direccion = form.get("direccion", "").strip()
    categoria = form.get("categoria", "Normal").strip()
    documento_identidad = form.get("documento_identidad", "").strip()
    if categoria not in CATEGORIAS_CLIENTE:
        categoria = "Normal"

    if email:
        cursor.execute(f"SELECT id FROM clientes WHERE email = {ph()} AND deleted_at IS NULL", (email,))
        if cursor.fetchone():
            return None, "Ya existe un cliente con ese email. Búscalo y edítalo."

    if documento_identidad:
        cursor.execute(f"SELECT id FROM clientes WHERE documento_identidad = {ph()} AND deleted_at IS NULL", (documento_identidad,))
        if cursor.fetchone():
            return None, "Ya existe un cliente con ese documento de identidad. Búscalo y edítalo."

    cursor.execute(f"""
        INSERT INTO clientes (nombre, empresa, contacto, telefono, email, direccion, categoria, documento_identidad)
        VALUES ({ph()}, {ph()}, {ph()}, {ph()}, {ph()}, {ph()}, {ph()}, {ph()})
    """, (nombre, empresa, contacto, telefono, email, direccion, categoria, documento_identidad or None))

    return cursor.lastrowid, None


@admin_bp.route("/clientes", methods=["GET", "POST"])
def admin_clientes():
    if not requiere_admin():
        return redirect("/login")

    conexion = conectar()
    cursor = conexion.cursor()

    error = None
    if request.method == "POST":
        cliente_id, error = _validar_y_crear_cliente(cursor, request.form)
        if not error:
            conexion.commit()
            conexion.close()
            return redirect("/admin/clientes?ok=1")
        # Si hubo error de validación, no se comitea nada (el INSERT no llegó a
        # correr) y se cae al listado de abajo conservando lo ya escrito.

    buscar_cl  = request.args.get("buscar", "").strip()
    filtro_cat = request.args.get("categoria", "").strip()
    pagina     = max(1, int(request.args.get("pagina", 1) or 1))
    por_pagina = 25

    cond_cl = ["deleted_at IS NULL"]
    params_cl = []
    if buscar_cl:
        cond_cl.append(f"(nombre LIKE {ph()} OR email LIKE {ph()} OR empresa LIKE {ph()} OR telefono LIKE {ph()})")
        like_cl = f"%{buscar_cl}%"
        params_cl.extend([like_cl, like_cl, like_cl, like_cl])
    if filtro_cat and filtro_cat in CATEGORIAS_CLIENTE:
        cond_cl.append(f"COALESCE(categoria, 'Normal') = {ph()}")
        params_cl.append(filtro_cat)
    where_cl = ("WHERE " + " AND ".join(cond_cl)) if cond_cl else ""

    cursor.execute(f"SELECT COUNT(*) AS total FROM clientes {where_cl}", params_cl)
    total_cl = cursor.fetchone()["total"]
    total_paginas_cl = max(1, (total_cl + por_pagina - 1) // por_pagina)
    pagina = min(pagina, total_paginas_cl)
    offset_cl = (pagina - 1) * por_pagina

    cursor.execute(f"""
        SELECT id, nombre, empresa, contacto, telefono, email, direccion,
               COALESCE(categoria, 'Normal') AS categoria, fecha_creacion
        FROM clientes
        {where_cl}
        ORDER BY id DESC
        LIMIT ? OFFSET ?
    """, params_cl + [por_pagina, offset_cl])
    lista = cursor.fetchall()

    conexion.close()

    return render_template("admin/clientes.html",
                           lista=lista,
                           buscar_cl=buscar_cl,
                           filtro_cat=filtro_cat,
                           categorias=CATEGORIAS_CLIENTE,
                           pagina_actual=pagina,
                           total_paginas=total_paginas_cl,
                           total=total_cl,
                           error=error,
                           form_data=request.form if error else None)


@admin_bp.route("/clientes/crear-rapido", methods=["POST"])
def crear_cliente_rapido():
    if not requiere_admin():
        return jsonify({"error": "No autorizado"}), 401

    conexion = conectar()
    cursor = conexion.cursor()

    cliente_id, error = _validar_y_crear_cliente(cursor, request.form)
    if error:
        conexion.close()
        return jsonify({"error": error}), 400

    conexion.commit()
    conexion.close()

    nombre = request.form.get("nombre", "").strip()
    email = request.form.get("email", "").strip()
    registrar_auditoria(f"Creó cliente {nombre} (rápido, desde alta de usuario)", "Clientes", "cliente", cliente_id)

    return jsonify({"ok": True, "id": cliente_id, "nombre": nombre, "email": email})


@admin_bp.route("/clientes/<int:id>/editar", methods=["GET", "POST"])
def editar_cliente(id):
    if not requiere_admin():
        return redirect("/login")

    conexion = conectar()
    cursor = conexion.cursor()

    if request.method == "POST":
        nombre = request.form["nombre"].strip()
        empresa = request.form.get("empresa", "").strip()
        contacto = request.form.get("contacto", "").strip()
        telefono = request.form.get("telefono", "").strip()
        email = request.form.get("email", "").strip()
        direccion = request.form.get("direccion", "").strip()
        categoria = request.form.get("categoria", "Normal").strip()
        documento_identidad = request.form.get("documento_identidad", "").strip()
        if categoria not in CATEGORIAS_CLIENTE:
            categoria = "Normal"

        if email:
            cursor.execute(f"SELECT id FROM clientes WHERE email = {ph()} AND deleted_at IS NULL AND id != {ph()}", (email, id))
            if cursor.fetchone():
                conexion.close()
                return redirect(f"/admin/clientes/{id}/editar?access_error=Ya+existe+otro+cliente+con+ese+email.")

        if documento_identidad:
            cursor.execute(f"SELECT id FROM clientes WHERE documento_identidad = {ph()} AND deleted_at IS NULL AND id != {ph()}", (documento_identidad, id))
            if cursor.fetchone():
                conexion.close()
                return redirect(f"/admin/clientes/{id}/editar?access_error=Ya+existe+otro+cliente+con+ese+documento.")

        cursor.execute(f"""
            UPDATE clientes
            SET nombre = {ph()}, empresa = {ph()}, contacto = {ph()}, telefono = {ph()},
                email = {ph()}, direccion = {ph()}, categoria = {ph()}, documento_identidad = {ph()}
            WHERE id = {ph()}
        """, (nombre, empresa, contacto, telefono, email, direccion, categoria, documento_identidad or None, id))

        conexion.commit()
        conexion.close()

        return redirect("/admin/clientes")

    cursor.execute(f"""
        SELECT id, nombre, empresa, contacto, telefono, email, direccion,
               COALESCE(categoria, 'Normal') AS categoria, documento_identidad
        FROM clientes
        WHERE id = {ph()}
    """, (id,))
    cliente = cursor.fetchone()

    conexion.close()

    if not cliente:
        return redirect("/admin/clientes")

    return render_template("admin/editar_cliente.html",
                           cliente=cliente,
                           categorias=CATEGORIAS_CLIENTE)


@admin_bp.route("/clientes/<int:id>/eliminar", methods=["POST"])
def eliminar_cliente(id):
    if session.get("rol") != "admin":
        return redirect("/login")

    conexion = conectar()
    cursor = conexion.cursor()
    cursor.execute(f"""
        UPDATE clientes SET deleted_at = CURRENT_TIMESTAMP, deleted_by = {ph()}
        WHERE id = {ph()}
    """, (session.get("usuario"), id))
    conexion.commit()
    conexion.close()
    registrar_auditoria("eliminó (papelera)", "clientes", "cliente", id)
    return redirect("/admin/clientes")


# ── Vehículos CRUD ──────────────────────────────────────────────────────────

@admin_bp.route("/vehiculos", methods=["GET", "POST"])
def admin_vehiculos():
    if not requiere_admin():
        return redirect("/login")

    conexion = conectar()
    cursor = conexion.cursor()

    if request.method == "POST":
        matricula = request.form["matricula"].strip()
        tipo = request.form.get("tipo", "").strip()
        marca = request.form.get("marca", "").strip()
        modelo = request.form.get("modelo", "").strip()
        capacidad = request.form.get("capacidad", "").strip()
        camionero_id = request.form.get("camionero_id") or None
        estado = request.form.get("estado", "Disponible").strip()

        cursor.execute(f"""
            INSERT INTO vehiculos (matricula, tipo, marca, modelo, capacidad, camionero_id, estado)
            VALUES ({ph()}, {ph()}, {ph()}, {ph()}, {ph()}, {ph()}, {ph()})
        """, (matricula, tipo, marca, modelo, capacidad, camionero_id, estado))
        conexion.commit()
        conexion.close()
        return redirect("/admin/vehiculos?ok=1")

    buscar_v = request.args.get("buscar", "").strip()
    filtro_estado_v = request.args.get("estado", "").strip()

    cond_v = []
    params_v = []
    if buscar_v:
        cond_v.append(f"(v.matricula LIKE {ph()} OR v.marca LIKE {ph()} OR v.modelo LIKE {ph()})")
        like_v = f"%{buscar_v}%"
        params_v.extend([like_v, like_v, like_v])
    if filtro_estado_v:
        cond_v.append(f"LOWER(v.estado) = {ph()}")
        params_v.append(filtro_estado_v.lower())
    where_v = "WHERE v.activo = 1" + (" AND " + " AND ".join(cond_v) if cond_v else "")

    cursor.execute(f"""
        SELECT v.id, v.matricula, v.tipo, v.marca, v.modelo, v.capacidad, v.estado,
               c.nombre AS camionero_nombre
        FROM vehiculos v
        LEFT JOIN camioneros c ON v.camionero_id = c.id
        {where_v}
        ORDER BY v.id DESC
    """, params_v)
    lista = cursor.fetchall()

    cursor.execute("SELECT id, nombre FROM camioneros WHERE activo = 1 ORDER BY nombre")
    camioneros = cursor.fetchall()

    conexion.close()

    return render_template(
        "admin/vehiculos.html",
        lista=lista,
        camioneros=camioneros,
        estados=VEHICULO_ESTADOS,
        buscar_v=buscar_v,
        filtro_estado_v=filtro_estado_v,
        estados_vehiculo=["Disponible", "En ruta", "En mantenimiento"],
    )


@admin_bp.route("/vehiculos/<int:id>/editar", methods=["GET", "POST"])
def editar_vehiculo(id):
    if not requiere_admin():
        return redirect("/login")

    conexion = conectar()
    cursor = conexion.cursor()

    if request.method == "POST":
        matricula = request.form["matricula"].strip()
        tipo = request.form.get("tipo", "").strip()
        marca = request.form.get("marca", "").strip()
        modelo = request.form.get("modelo", "").strip()
        capacidad = request.form.get("capacidad", "").strip()
        camionero_id = request.form.get("camionero_id") or None
        estado = request.form.get("estado", "Disponible").strip()

        cursor.execute(f"""
            UPDATE vehiculos
            SET matricula = {ph()}, tipo = {ph()}, marca = {ph()}, modelo = {ph()},
                capacidad = {ph()}, camionero_id = {ph()}, estado = {ph()}
            WHERE id = {ph()}
        """, (matricula, tipo, marca, modelo, capacidad, camionero_id, estado, id))
        conexion.commit()
        conexion.close()
        return redirect("/admin/vehiculos")

    cursor.execute(f"""
        SELECT id, matricula, tipo, marca, modelo, capacidad, camionero_id, estado
        FROM vehiculos
        WHERE id = {ph()} AND activo = 1
    """, (id,))
    vehiculo = cursor.fetchone()

    cursor.execute("SELECT id, nombre FROM camioneros WHERE activo = 1 ORDER BY nombre")
    camioneros = cursor.fetchall()

    conexion.close()

    if not vehiculo:
        return redirect("/admin/vehiculos")

    return render_template(
        "admin/editar_vehiculo.html",
        vehiculo=vehiculo,
        camioneros=camioneros,
        estados=VEHICULO_ESTADOS
    )


@admin_bp.route("/vehiculos/sugerencias")
def sugerencias_vehiculos():
    if not requiere_admin():
        return jsonify([]), 403

    campo = request.args.get("campo", "")
    q = request.args.get("q", "").strip()

    if campo not in ("marca", "modelo"):
        return jsonify([]), 400

    conexion = conectar()
    cursor = conexion.cursor()
    cursor.execute(
        f"SELECT DISTINCT {campo} FROM vehiculos"
        f" WHERE {campo} LIKE {ph()} AND activo = 1 AND {campo} != ''"
        f" ORDER BY {campo} LIMIT 10",
        (f"%{q}%",)
    )
    resultados = [row[0] for row in cursor.fetchall() if row[0]]
    conexion.close()

    return jsonify(resultados)


@admin_bp.route("/vehiculos/<int:id>/eliminar", methods=["POST"])
def eliminar_vehiculo(id):
    if session.get("rol") != "admin":
        return redirect("/login")

    conexion = conectar()
    cursor = conexion.cursor()
    cursor.execute(f"UPDATE vehiculos SET activo = 0 WHERE id = {ph()}", (id,))
    conexion.commit()
    conexion.close()

    return redirect("/admin/vehiculos")


# ── Reportes ─────────────────────────────────────────────────────────────────

MESES_ES = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
            "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]


def _ultimos_6_meses():
    """Devuelve lista de (year, month) para los últimos 6 meses, orden ascendente."""
    hoy = date.today()
    meses = []
    for i in range(5, -1, -1):
        m = hoy.month - i
        y = hoy.year
        while m <= 0:
            m += 12
            y -= 1
        meses.append((y, m))
    return meses


def _calcular_financieros_periodo(fecha_desde, fecha_hasta,
                                   estado_cobro=None, forma_pago=None):
    """Devuelve (filas_tabla, totales) para el período dado con filtros opcionales de cobro."""
    conexion = conectar()
    cursor = conexion.cursor()

    where_extra = ""
    params = [fecha_desde, fecha_hasta + " 23:59:59"]
    if estado_cobro == "Cobrado":
        where_extra += f" AND fecha_cobro IS NOT NULL"
    elif estado_cobro == "Pendiente":
        where_extra += f" AND fecha_cobro IS NULL"
    if forma_pago:
        where_extra += f" AND forma_cobro = {ph()}"
        params.append(forma_pago)

    cursor.execute(f"""
        SELECT id, cliente, origen, destino, estado, fecha_creacion,
               forma_cobro, codigo_transaccion, fecha_cobro, monto_cobrado,
               verificado_financiero, verificado_por, fecha_verificacion
        FROM viajes
        WHERE (fecha_creacion >= {ph()} AND fecha_creacion <= {ph()})
          AND LOWER(estado) != 'cancelado'
          {where_extra}
        ORDER BY id DESC
    """, params)
    viajes_periodo = cursor.fetchall()
    conexion.close()

    filas = []
    totales = {"ingresos": 0.0, "pago_camionero": 0.0,
               "combustible": 0.0, "comision": 0.0, "utilidad": 0.0}

    for v in viajes_periodo:
        liq = calcular_liquidacion(v["id"])
        pc = liq["precio_cliente"] if liq else 0.0
        pag = liq["pago_camionero"] if liq else 0.0
        comb = liq["combustible"] if liq else 0.0
        com = liq["comision_mercatoria"] if liq else 0.0
        util = liq["utilidad_mercatoria"] if liq else 0.0

        totales["ingresos"] += pc
        totales["pago_camionero"] += pag
        totales["combustible"] += comb
        totales["comision"] += com
        totales["utilidad"] += util

        filas.append({
            "id": v["id"],
            "cliente": v["cliente"] or "—",
            "ruta": f"{v['origen']} → {v['destino']}",
            "estado": v["estado"],
            "precio_cliente": pc,
            "pago_camionero": pag,
            "combustible": comb,
            "comision": com,
            "utilidad": util,
            "cobrado": bool(v["fecha_cobro"]),
            "forma_cobro": v["forma_cobro"] or "",
            "codigo_transaccion": v["codigo_transaccion"] or "",
            "fecha_cobro": v["fecha_cobro"] or "",
            "monto_cobrado": v["monto_cobrado"],
            "verificado": bool(v["verificado_financiero"]),
            "verificado_por": v["verificado_por"] or "",
            "fecha_verificacion": v["fecha_verificacion"] or "",
        })

    return filas, totales


@admin_bp.route("/reportes")
def reportes():
    if not (session.get("usuario") and session.get("rol") == "admin"):
        return redirect("/admin")

    hoy = date.today()
    fecha_desde = request.args.get("fecha_desde", hoy.replace(day=1).isoformat())
    fecha_hasta = request.args.get("fecha_hasta", hoy.isoformat())
    estado_cobro = request.args.get("estado_cobro", "")
    forma_pago = request.args.get("forma_pago", "")

    filas_tabla, totales = _calcular_financieros_periodo(
        fecha_desde, fecha_hasta,
        estado_cobro=estado_cobro or None,
        forma_pago=forma_pago or None,
    )

    # KPIs adicionales
    n = len(filas_tabla)
    viaje_promedio = totales["ingresos"] / n if n else 0.0

    conexion = conectar()
    cursor = conexion.cursor()

    cursor.execute(f"""
        SELECT origen || ' → ' || destino AS ruta,
               COUNT(*) AS viajes,
               SUM(COALESCE(NULLIF(precio_final,0), NULLIF(precio_cliente,0), NULLIF(precio,0), 0)) AS ingresos_proxy
        FROM viajes
        WHERE (fecha_creacion >= {ph()} AND fecha_creacion <= {ph()})
          AND LOWER(estado) != 'cancelado'
        GROUP BY ruta
        ORDER BY ingresos_proxy DESC
        LIMIT 1
    """, (fecha_desde, fecha_hasta + " 23:59:59"))
    ruta_top = cursor.fetchone()

    cursor.execute(f"""
        SELECT camionero_nombre, COUNT(*) AS total_viajes
        FROM viajes
        WHERE (fecha_creacion >= {ph()} AND fecha_creacion <= {ph()})
          AND LOWER(estado) != 'cancelado'
          AND camionero_nombre IS NOT NULL AND camionero_nombre != ''
        GROUP BY camionero_nombre
        ORDER BY total_viajes DESC
        LIMIT 1
    """, (fecha_desde, fecha_hasta + " 23:59:59"))
    camionero_top = cursor.fetchone()

    # Datos mensuales (últimos 6 meses) para el gráfico
    if USE_POSTGRES:
        cursor.execute("""
            SELECT TO_CHAR(fecha_creacion, 'YYYY-MM') AS mes,
                   SUM(COALESCE(NULLIF(precio_final,0), NULLIF(precio_cliente,0), NULLIF(precio,0), 0)) AS ingresos
            FROM viajes
            WHERE fecha_creacion >= DATE_TRUNC('month', CURRENT_DATE - INTERVAL '5 months')
              AND LOWER(estado) != 'cancelado'
            GROUP BY mes
            ORDER BY mes
        """)
    else:
        cursor.execute("""
            SELECT strftime('%Y-%m', fecha_creacion) AS mes,
                   SUM(COALESCE(NULLIF(precio_final,0), NULLIF(precio_cliente,0), NULLIF(precio,0), 0)) AS ingresos
            FROM viajes
            WHERE fecha_creacion >= date('now', '-5 months', 'start of month')
              AND LOWER(estado) != 'cancelado'
            GROUP BY mes
            ORDER BY mes
        """)
    datos_db = {row["mes"]: float(row["ingresos"] or 0) for row in cursor.fetchall()}
    conexion.close()

    # Rellenar los 6 meses aunque no haya datos
    meses_base = _ultimos_6_meses()
    chart_labels = [f"{MESES_ES[m-1]} {y}" for y, m in meses_base]
    chart_data = [datos_db.get(f"{y:04d}-{m:02d}", 0.0) for y, m in meses_base]

    return render_template(
        "admin/reportes.html",
        filas_tabla=filas_tabla,
        totales=totales,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        estado_cobro=estado_cobro,
        forma_pago=forma_pago,
        chart_labels=json.dumps(chart_labels),
        chart_data=json.dumps(chart_data),
        viaje_promedio=round(viaje_promedio, 2),
        ruta_top=ruta_top,
        camionero_top=camionero_top,
        total_viajes=n,
    )


@admin_bp.route("/reportes/exportar")
def exportar_reportes_csv():
    if not (session.get("usuario") and session.get("rol") == "admin"):
        return redirect("/admin?access_error=Solo+administradores+pueden+ver+reportes")

    hoy = date.today()
    fecha_desde = request.args.get("fecha_desde", hoy.replace(day=1).isoformat())
    fecha_hasta = request.args.get("fecha_hasta", hoy.isoformat())

    filas, totales = _calcular_financieros_periodo(fecha_desde, fecha_hasta)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Cliente", "Ruta", "Estado",
                     "Precio Cliente (USD)", "Pago Camionero (USD)",
                     "Combustible (USD)", "Utilidad (USD)"])
    for f in filas:
        writer.writerow([
            f["id"], f["cliente"], f["ruta"], f["estado"],
            f"{f['precio_cliente']:.2f}", f"{f['pago_camionero']:.2f}",
            f"{f['combustible']:.2f}", f"{f['utilidad']:.2f}",
        ])
    writer.writerow([])
    writer.writerow(["TOTALES", "", "", "",
                     f"{totales['ingresos']:.2f}", f"{totales['pago_camionero']:.2f}",
                     f"{totales['combustible']:.2f}", f"{totales['utilidad']:.2f}"])

    csv_bytes = output.getvalue().encode("utf-8-sig")
    nombre = f"reporte-{fecha_desde}-{fecha_hasta}.csv"
    return send_file(
        io.BytesIO(csv_bytes),
        mimetype="text/csv",
        as_attachment=True,
        download_name=nombre,
    )


# ── Usuarios CRUD ─────────────────────────────────────────────────────────────

@admin_bp.route("/usuarios", methods=["GET"])
def lista_usuarios():
    if session.get("rol") != "admin":
        return redirect("/login")

    filtro_rol = request.args.get("rol", "").strip().lower()

    conexion = conectar()
    cursor = conexion.cursor()

    if filtro_rol:
        cursor.execute(f"""
            SELECT id, usuario, rol, activo, fecha_creacion
            FROM usuarios
            WHERE rol = {ph()}
            ORDER BY id DESC
        """, (filtro_rol,))
    else:
        cursor.execute("""
            SELECT id, usuario, rol, activo, fecha_creacion
            FROM usuarios
            ORDER BY id DESC
        """)
    lista = cursor.fetchall()

    cursor.execute("SELECT COUNT(*) AS total FROM usuarios")
    total = cursor.fetchone()["total"]
    cursor.execute("SELECT COUNT(*) AS total FROM usuarios WHERE rol = 'admin'")
    total_admin = cursor.fetchone()["total"]
    cursor.execute("SELECT COUNT(*) AS total FROM usuarios WHERE rol = 'operador'")
    total_operador = cursor.fetchone()["total"]
    cursor.execute("SELECT COUNT(*) AS total FROM usuarios WHERE rol = 'cliente'")
    total_cliente = cursor.fetchone()["total"]

    cursor.execute("""
        SELECT id, nombre, email
        FROM clientes
        WHERE usuario_id IS NULL AND deleted_at IS NULL
        ORDER BY nombre
    """)
    clientes_sin_usuario = cursor.fetchall()

    conexion.close()

    return render_template(
        "admin/usuarios.html",
        usuarios=lista,
        filtro_rol=filtro_rol,
        total=total,
        total_admin=total_admin,
        total_operador=total_operador,
        total_cliente=total_cliente,
        clientes_sin_usuario=clientes_sin_usuario,
    )


@admin_bp.route("/usuarios/crear", methods=["POST"])
def crear_usuario():
    if session.get("rol") != "admin":
        return redirect("/admin")

    usuario = request.form.get("usuario", "").strip()
    password = request.form.get("password", "").strip()
    rol = request.form.get("rol", "").strip()
    cliente_id_str = request.form.get("cliente_id", "").strip()

    if not usuario or not password or rol not in ("admin", "operador", "cliente"):
        return redirect("/admin/usuarios?error=Datos+inv%C3%A1lidos")

    if rol == "cliente" and not cliente_id_str:
        return redirect("/admin/usuarios?error=Selecciona+el+cliente+a+vincular")

    cliente_id = None
    if rol == "cliente":
        try:
            cliente_id = int(cliente_id_str)
        except ValueError:
            return redirect("/admin/usuarios?error=Cliente+inv%C3%A1lido")

    conexion = conectar()
    cursor = conexion.cursor()
    cursor.execute(f"SELECT id FROM usuarios WHERE usuario = {ph()}", (usuario,))
    if cursor.fetchone():
        conexion.close()
        return redirect("/admin/usuarios?error=El+usuario+ya+existe")

    if cliente_id is not None:
        cursor.execute(f"SELECT id FROM clientes WHERE id = {ph()} AND deleted_at IS NULL", (cliente_id,))
        if not cursor.fetchone():
            conexion.close()
            return redirect("/admin/usuarios?error=El+cliente+seleccionado+no+existe")

    hash_pw = bcrypt.generate_password_hash(password).decode("utf-8")
    cursor.execute(
        f"INSERT INTO usuarios (usuario, password, rol) VALUES ({ph()}, {ph()}, {ph()})",
        (usuario, hash_pw, rol)
    )
    nuevo_usuario_id = cursor.lastrowid

    if cliente_id is not None:
        cursor.execute(
            f"UPDATE clientes SET usuario_id = {ph()} WHERE id = {ph()} AND usuario_id IS NULL",
            (nuevo_usuario_id, cliente_id)
        )
        if cursor.rowcount == 0:
            # Otro admin vinculó este cliente en el intervalo: no comiteamos nada,
            # ni el usuario ni el update quedan persistidos (transacción atómica).
            conexion.close()
            return redirect("/admin/usuarios?error=Ese+cliente+ya+fue+vinculado+por+otro+usuario%2C+intenta+de+nuevo")

    conexion.commit()
    conexion.close()

    detalle = f"Creó usuario {usuario} con rol {rol}"
    if cliente_id is not None:
        detalle += f", vinculado a cliente #{cliente_id}"
    registrar_auditoria(detalle, "Usuarios", "usuario", nuevo_usuario_id)

    return redirect("/admin/usuarios")


@admin_bp.route("/usuarios/<int:id>/rol", methods=["POST"])
def cambiar_rol_usuario(id):
    if session.get("rol") != "admin":
        return redirect("/admin")

    rol = request.form.get("rol", "").strip()
    if rol not in ("admin", "operador", "cliente"):
        return redirect("/admin/usuarios?error=Rol+inv%C3%A1lido")

    conexion = conectar()
    cursor = conexion.cursor()
    cursor.execute(f"SELECT usuario FROM usuarios WHERE id = {ph()}", (id,))
    row = cursor.fetchone()
    if row and row["usuario"] == session.get("usuario"):
        conexion.close()
        return redirect("/admin/usuarios?error=No+puedes+cambiar+tu+propio+rol")

    cursor.execute(f"UPDATE usuarios SET rol = {ph()} WHERE id = {ph()}", (rol, id))
    conexion.commit()
    conexion.close()

    registrar_auditoria(f"Cambió rol de usuario #{id} a {rol}", "Usuarios", "usuario", id)

    return redirect("/admin/usuarios")


@admin_bp.route("/usuarios/<int:id>/toggle", methods=["POST"])
def toggle_usuario(id):
    if session.get("rol") != "admin":
        return redirect("/admin")

    conexion = conectar()
    cursor = conexion.cursor()
    cursor.execute(f"SELECT usuario FROM usuarios WHERE id = {ph()}", (id,))
    row = cursor.fetchone()
    if row and row["usuario"] == session.get("usuario"):
        conexion.close()
        return redirect("/admin/usuarios?error=No+puedes+desactivarte+a+ti+mismo")

    cursor.execute(
        f"UPDATE usuarios SET activo = CASE WHEN activo = 1 THEN 0 ELSE 1 END WHERE id = {ph()}",
        (id,)
    )
    conexion.commit()
    conexion.close()

    registrar_auditoria(f"Cambió estado de usuario #{id}", "Usuarios", "usuario", id)

    return redirect("/admin/usuarios")


@admin_bp.route("/usuarios/<int:id>/reset-password", methods=["POST"])
def reset_password_usuario(id):
    if session.get("rol") != "admin":
        return redirect("/admin/usuarios")
    nueva = request.form.get("nueva_password", "").strip()
    if len(nueva) < 4:
        return redirect("/admin/usuarios?error=La+contraseña+debe+tener+al+menos+4+caracteres")
    from extensions import bcrypt as _bcrypt
    nuevo_hash = _bcrypt.generate_password_hash(nueva).decode("utf-8")
    conexion = conectar()
    cursor = conexion.cursor()
    cursor.execute(f"UPDATE usuarios SET password = {ph()} WHERE id = {ph()}", (nuevo_hash, id))
    conexion.commit()
    conexion.close()
    registrar_auditoria(f"Reseteó contraseña de usuario #{id}", "Usuarios", "usuario", id)
    return redirect("/admin/usuarios?ok=Contraseña+actualizada")


@admin_bp.route("/mi-cuenta", methods=["GET", "POST"])
def mi_cuenta():
    if not requiere_admin():
        return redirect("/login")

    conexion = conectar()
    cursor = conexion.cursor()
    cursor.execute(f"SELECT id, password FROM usuarios WHERE usuario = {ph()}", (session["usuario"],))
    fila = cursor.fetchone()

    if request.method == "POST":
        actual = request.form.get("actual", "").strip()
        nueva = request.form.get("nueva", "").strip()
        confirmar = request.form.get("confirmar", "").strip()

        if not fila or not bcrypt.check_password_hash(fila["password"], actual):
            conexion.close()
            return render_template("admin/mi_cuenta.html", error="La contraseña actual es incorrecta")

        if nueva != confirmar:
            conexion.close()
            return render_template("admin/mi_cuenta.html", error="Las contraseñas nuevas no coinciden")

        if len(nueva) < 4:
            conexion.close()
            return render_template("admin/mi_cuenta.html", error="La nueva contraseña debe tener al menos 4 caracteres")

        nuevo_hash = bcrypt.generate_password_hash(nueva).decode("utf-8")
        cursor.execute(f"UPDATE usuarios SET password = {ph()} WHERE usuario = {ph()}", (nuevo_hash, session["usuario"]))
        conexion.commit()
        conexion.close()
        registrar_auditoria("Cambió su propia contraseña", "Usuarios", "usuario", fila["id"])
        return render_template("admin/mi_cuenta.html", mensaje="Contraseña actualizada correctamente")

    conexion.close()
    return render_template("admin/mi_cuenta.html")


# ── Exportar / Importar Excel ──────────────────────────────────────────────────

_EXCEL_CONFIG = {
    "rutas": {
        "columnas": ["id", "nombre", "origen", "destino", "zona", "km_oficiales"],
        "tabla": "rutas",
        "redirect": "/admin/comercial/rutas",
    },
    "camioneros": {
        "columnas": ["id", "nombre", "telefono", "licencia", "estado"],
        "tabla": "camioneros",
        "redirect": "/admin/camioneros",
    },
    "clientes": {
        "columnas": ["id", "nombre", "contacto", "telefono", "email", "empresa"],
        "tabla": "clientes",
        "redirect": "/admin/clientes",
    },
    "vehiculos": {
        "columnas": ["id", "nombre", "marca", "modelo", "matricula", "tipo", "capacidad", "estado"],
        "tabla": "vehiculos",
        "redirect": "/admin/vehiculos",
    },
}

_HEADER_FILL = PatternFill("solid", fgColor="E86A2C")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


@admin_bp.route("/exportar/<string:tabla>", methods=["GET"])
def exportar_excel(tabla):
    if session.get("rol") != "admin":
        return redirect("/admin")

    cfg = _EXCEL_CONFIG.get(tabla)
    if not cfg:
        return redirect("/admin"), 404

    conexion = conectar()
    cursor = conexion.cursor()
    cols = ", ".join(cfg["columnas"])
    cursor.execute(f"SELECT {cols} FROM {cfg['tabla']}")
    filas = cursor.fetchall()
    conexion.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = tabla.capitalize()

    for col_idx, col_name in enumerate(cfg["columnas"], start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name.upper())
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row_idx, fila in enumerate(filas, start=2):
        for col_idx, valor in enumerate(fila, start=1):
            ws.cell(row=row_idx, column=col_idx, value=valor)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    registrar_auditoria(f"Exportó Excel de {tabla}", "Datos", tabla)

    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"{tabla}.xlsx",
    )


@admin_bp.route("/importar/<string:tabla>", methods=["POST"])
def importar_excel(tabla):
    if session.get("rol") != "admin":
        return redirect("/admin")

    cfg = _EXCEL_CONFIG.get(tabla)
    if not cfg:
        return redirect("/admin"), 404

    archivo = request.files.get("archivo")
    if not archivo:
        return redirect(cfg["redirect"] + "?error=No+se+recibió+archivo")

    wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
    ws = wb.active

    columnas = cfg["columnas"]
    placeholders = ", ".join([f"{ph()}"] * len(columnas))
    cols_str = ", ".join(columnas)

    conexion = conectar()
    cursor = conexion.cursor()
    importados = 0
    errores = []

    for idx, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        valores = list(fila[: len(columnas)])
        if not any(v is not None for v in valores):
            continue
        try:
            cursor.execute(
                f"INSERT INTO {cfg['tabla']} ({cols_str}) VALUES ({placeholders}) "
                "ON CONFLICT (id) DO NOTHING",
                valores,
            )
            importados += cursor.rowcount
        except Exception as e:
            errores.append((idx, str(e)))

    conexion.commit()
    conexion.close()

    if errores:
        from flask import current_app
        detalle_errores = "; ".join(f"fila {idx}: {msg}" for idx, msg in errores)
        current_app.logger.error(f"Importación Excel a {tabla} con {len(errores)} fila(s) fallida(s): {detalle_errores}")

    registrar_auditoria(f"Importó Excel a {tabla} ({importados} registros, {len(errores)} fallidas)", "Datos", tabla)

    redirect_url = f"{cfg['redirect']}?importado={importados}+registros"
    if errores:
        redirect_url += f"&fallidos={len(errores)}"
    return redirect(redirect_url)


# ── Auditoría ──────────────────────────────────────────────────────────────────

@admin_bp.route("/auditoria", methods=["GET"])
def ver_auditoria():
    if session.get("rol") != "admin":
        return redirect("/admin")

    categoria = request.args.get("categoria", "").strip()
    usuario_f = request.args.get("usuario", "").strip()
    fecha_desde = request.args.get("fecha_desde", "").strip()
    fecha_hasta = request.args.get("fecha_hasta", "").strip()
    buscar = request.args.get("buscar", "").strip()
    pagina = max(1, int(request.args.get("pagina", 1) or 1))
    por_pagina = 50

    condiciones = []
    params = []

    if categoria:
        condiciones.append(f"categoria = {ph()}")
        params.append(categoria)
    if usuario_f:
        condiciones.append(f"usuario LIKE {ph()}")
        params.append(f"%{usuario_f}%")
    if fecha_desde:
        condiciones.append(f"DATE(fecha) >= {ph()}")
        params.append(fecha_desde)
    if fecha_hasta:
        condiciones.append(f"DATE(fecha) <= {ph()}")
        params.append(fecha_hasta)
    if buscar:
        condiciones.append(f"(accion LIKE {ph()} OR detalle LIKE {ph()} OR usuario LIKE {ph()})")
        params.extend([f"%{buscar}%", f"%{buscar}%", f"%{buscar}%"])

    where = ("WHERE " + " AND ".join(condiciones)) if condiciones else ""

    conexion = conectar()
    cursor = conexion.cursor()

    cursor.execute(f"SELECT COUNT(*) as total FROM auditoria {where}", params)
    total = cursor.fetchone()["total"]

    offset = (pagina - 1) * por_pagina
    cursor.execute(
        f"SELECT * FROM auditoria {where} ORDER BY id DESC LIMIT {ph()} OFFSET {ph()}",
        params + [por_pagina, offset]
    )
    registros = cursor.fetchall()
    conexion.close()

    total_paginas = (total + por_pagina - 1) // por_pagina

    return render_template(
        "admin/auditoria.html",
        registros=registros,
        pagina=pagina,
        total_paginas=total_paginas,
        total=total,
        filtros=dict(
            categoria=categoria,
            usuario=usuario_f,
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
            buscar=buscar,
        ),
    )


# ── Acciones por lote ──────────────────────────────────────────────────────────

_ESTADOS_VIAJE = [
    "Solicitado", "Pendiente", "Asignado", "En ruta",
    "Carga recogida", "Entregado", "Cancelado",
]


@admin_bp.route("/lote", methods=["GET"])
def panel_lote():
    if session.get("rol") != "admin":
        return redirect("/admin")

    conexion = conectar()
    cursor = conexion.cursor()
    cursor.execute("""
        SELECT id, nombre, origen, destino
        FROM rutas ORDER BY COALESCE(nombre, origen)
    """)
    rutas = cursor.fetchall()
    cursor.execute("""
        SELECT id, cliente, origen, destino, estado,
               COALESCE(NULLIF(precio_cliente,0), NULLIF(precio_final,0), NULLIF(precio,0), 0) AS precio
        FROM viajes
        WHERE LOWER(estado) NOT IN ('entregado', 'cancelado')
        ORDER BY id DESC
    """)
    viajes_activos = cursor.fetchall()
    conexion.close()

    resultado = request.args.get("resultado", "")
    return render_template(
        "admin/lote.html",
        rutas=rutas,
        viajes_activos=viajes_activos,
        estados=_ESTADOS_VIAJE,
        resultado=resultado,
    )


@admin_bp.route("/lote/preview", methods=["GET"])
def lote_preview():
    if session.get("rol") != "admin":
        return jsonify({"count": 0})

    criterio = request.args.get("criterio", "")
    valor = request.args.get("valor", "").strip()

    conexion = conectar()
    cursor = conexion.cursor()

    try:
        if criterio == "sin_precio":
            cursor.execute("""
                SELECT COUNT(*) AS total FROM viajes
                WHERE LOWER(estado) NOT IN ('entregado','cancelado')
                  AND (precio_cliente IS NULL OR precio_cliente = 0)
                  AND (precio_final   IS NULL OR precio_final   = 0)
                  AND (precio         IS NULL OR precio         = 0)
            """)
        elif criterio == "estado" and valor:
            cursor.execute(
                f"SELECT COUNT(*) AS total FROM viajes WHERE LOWER(estado) = LOWER({ph()})", (valor,)
            )
        elif criterio == "ruta" and valor:
            cursor.execute(
                f"SELECT COUNT(*) AS total FROM viajes WHERE ruta_id = {ph()}", (valor,)
            )
        else:
            cursor.execute(
                "SELECT COUNT(*) AS total FROM viajes WHERE LOWER(estado) NOT IN ('entregado','cancelado')"
            )
        count = cursor.fetchone()["total"]
    except Exception:
        count = 0
    finally:
        conexion.close()

    return jsonify({"count": count})


@admin_bp.route("/lote/precios", methods=["POST"])
def lote_precios():
    if session.get("rol") != "admin":
        return redirect("/admin")

    criterio = request.form.get("criterio", "").strip()
    valor_criterio = request.form.get("valor_criterio", "").strip()
    motivo = request.form.get("motivo", "").strip()

    try:
        precio_nuevo = float(request.form.get("precio_nuevo", 0))
        if precio_nuevo <= 0:
            raise ValueError
    except (ValueError, TypeError):
        return redirect("/admin/lote?resultado=Error:+precio+inválido")

    conexion = conectar()
    cursor = conexion.cursor()

    if criterio == "sin_precio":
        cursor.execute(f"""
            UPDATE viajes SET precio_cliente = {ph()}, precio_final = {ph()}
            WHERE LOWER(estado) NOT IN ('entregado','cancelado')
              AND (precio_cliente IS NULL OR precio_cliente = 0)
              AND (precio_final   IS NULL OR precio_final   = 0)
              AND (precio         IS NULL OR precio         = 0)
        """, (precio_nuevo, precio_nuevo))
    elif criterio == "estado" and valor_criterio:
        cursor.execute(f"""
            UPDATE viajes SET precio_cliente = {ph()}, precio_final = {ph()}
            WHERE LOWER(estado) = LOWER({ph()})
        """, (precio_nuevo, precio_nuevo, valor_criterio))
    elif criterio == "ruta" and valor_criterio:
        cursor.execute(f"""
            UPDATE viajes SET precio_cliente = {ph()}, precio_final = {ph()}
            WHERE ruta_id = {ph()}
        """, (precio_nuevo, precio_nuevo, valor_criterio))
    else:
        cursor.execute(f"""
            UPDATE viajes SET precio_cliente = {ph()}, precio_final = {ph()}
            WHERE LOWER(estado) NOT IN ('entregado','cancelado')
        """, (precio_nuevo, precio_nuevo))

    n = cursor.rowcount
    conexion.commit()
    conexion.close()

    registrar_auditoria(
        f"Precio masivo ${precio_nuevo} aplicado a {n} viajes. Motivo: {motivo}",
        "Viajes", "viajes"
    )
    from urllib.parse import quote_plus as qp
    return redirect(f"/admin/lote?resultado={qp(str(n) + ' viajes actualizados con precio $' + str(precio_nuevo))}")


@admin_bp.route("/lote/estados", methods=["POST"])
def lote_estados():
    if session.get("rol") != "admin":
        return redirect("/admin")

    estado_origen = request.form.get("estado_origen", "").strip()
    estado_destino = request.form.get("estado_destino", "").strip()
    motivo = request.form.get("motivo", "").strip()

    if not estado_origen or not estado_destino:
        return redirect("/admin/lote?resultado=Error:+faltan+estados")

    conexion = conectar()
    cursor = conexion.cursor()
    cursor.execute(
        f"UPDATE viajes SET estado = {ph()} WHERE LOWER(estado) = LOWER({ph()})",
        (estado_destino, estado_origen)
    )
    n = cursor.rowcount
    conexion.commit()
    conexion.close()

    registrar_auditoria(
        f"Cambio masivo de estado: {estado_origen} → {estado_destino} ({n} viajes). Motivo: {motivo}",
        "Viajes", "viajes"
    )
    from urllib.parse import quote_plus as qp
    return redirect(f"/admin/lote?resultado={qp(str(n) + ' viajes cambiados a ' + estado_destino)}")


@admin_bp.route("/lote/viajes-seleccionados", methods=["POST"])
def lote_seleccionados():
    if session.get("rol") != "admin":
        return redirect("/admin")

    ids_raw = request.form.getlist("ids[]")
    accion = request.form.get("accion", "").strip()
    valor = request.form.get("valor", "").strip()
    motivo = request.form.get("motivo", "").strip()

    try:
        ids = [int(i) for i in ids_raw if i.isdigit()]
    except Exception:
        ids = []

    if not ids:
        return redirect("/admin/lote?resultado=Error:+no+se+seleccionaron+viajes")

    placeholders = ",".join(f"{ph()}" * len(ids))
    conexion = conectar()
    cursor = conexion.cursor()

    if accion == "precio":
        try:
            precio = float(valor)
        except (ValueError, TypeError):
            conexion.close()
            return redirect("/admin/lote?resultado=Error:+precio+inválido")
        cursor.execute(
            f"UPDATE viajes SET precio_cliente = {ph()}, precio_final = {ph()} WHERE id IN ({placeholders})",
            [precio, precio] + ids
        )
        detalle = f"precio ${precio}"
    elif accion == "estado":
        cursor.execute(
            f"UPDATE viajes SET estado = {ph()} WHERE id IN ({placeholders})",
            [valor] + ids
        )
        detalle = f"estado '{valor}'"
    else:
        conexion.close()
        return redirect("/admin/lote?resultado=Error:+acción+desconocida")

    n = cursor.rowcount
    conexion.commit()
    conexion.close()

    registrar_auditoria(
        f"Acción manual por lote sobre {n} viajes: {detalle}. Motivo: {motivo}",
        "Viajes", "viajes"
    )
    from urllib.parse import quote_plus as qp
    return redirect(f"/admin/lote?resultado={qp(str(n) + ' viajes actualizados (' + detalle + ')')}")


# ── Mensajes masivos ───────────────────────────────────────────────────────────

def _query_emails_clientes(filtro):
    conexion = conectar()
    cursor = conexion.cursor()

    if filtro == "activos":
        cursor.execute("""
            SELECT DISTINCT cl.email FROM clientes cl
            JOIN viajes v ON v.cliente_id = cl.id
            WHERE LOWER(v.estado) NOT IN ('entregado', 'cancelado')
              AND cl.email IS NOT NULL AND cl.email != ''
        """)
    elif filtro == "zona_occidente":
        cursor.execute("""
            SELECT DISTINCT cl.email FROM clientes cl
            JOIN viajes v ON v.cliente_id = cl.id
            JOIN rutas r ON v.ruta_id = r.id
            WHERE LOWER(r.zona) = 'occidente'
              AND cl.email IS NOT NULL AND cl.email != ''
        """)
    elif filtro == "zona_centro":
        cursor.execute("""
            SELECT DISTINCT cl.email FROM clientes cl
            JOIN viajes v ON v.cliente_id = cl.id
            JOIN rutas r ON v.ruta_id = r.id
            WHERE LOWER(r.zona) = 'centro'
              AND cl.email IS NOT NULL AND cl.email != ''
        """)
    elif filtro == "zona_oriente":
        cursor.execute("""
            SELECT DISTINCT cl.email FROM clientes cl
            JOIN viajes v ON v.cliente_id = cl.id
            JOIN rutas r ON v.ruta_id = r.id
            WHERE LOWER(r.zona) = 'oriente'
              AND cl.email IS NOT NULL AND cl.email != ''
        """)
    elif filtro == "sin_viajes":
        cursor.execute("""
            SELECT cl.email FROM clientes cl
            WHERE cl.email IS NOT NULL AND cl.email != ''
              AND cl.id NOT IN (
                  SELECT DISTINCT cliente_id FROM viajes WHERE cliente_id IS NOT NULL
              )
        """)
    else:
        cursor.execute("SELECT email FROM clientes WHERE email IS NOT NULL AND email != ''")

    emails = [row["email"] for row in cursor.fetchall() if row["email"]]
    conexion.close()
    return emails


@admin_bp.route("/mensajes")
def mensajes():
    if session.get("rol") != "admin":
        return redirect("/login")

    conexion = conectar()
    cursor = conexion.cursor()
    cursor.execute("""
        SELECT accion, fecha FROM auditoria
        WHERE accion LIKE 'Envió mensaje masivo%'
        ORDER BY id DESC
        LIMIT 10
    """)
    historial = cursor.fetchall()
    conexion.close()

    enviado = request.args.get("enviado")
    error = request.args.get("error")
    return render_template("admin/mensajes.html", historial=historial, enviado=enviado, error=error)


@admin_bp.route("/mensajes/preview")
def mensajes_preview():
    if session.get("rol") != "admin":
        return jsonify({"count": 0, "emails": []})

    filtro = request.args.get("filtro", "todos")
    emails = _query_emails_clientes(filtro)
    return jsonify({"count": len(emails), "emails": emails[:3]})


@admin_bp.route("/mensajes/enviar", methods=["POST"])
def mensajes_enviar():
    if session.get("rol") != "admin":
        return redirect("/login")

    asunto = request.form.get("asunto", "").strip()
    cuerpo = request.form.get("cuerpo", "").strip()
    destinatarios = request.form.get("destinatarios", "todos").strip()

    if not asunto or not cuerpo:
        return redirect("/admin/mensajes?error=Faltan+asunto+o+cuerpo")

    emails = _query_emails_clientes(destinatarios)

    from flask import current_app
    app_obj = current_app._get_current_object()

    def enviar_todos():
        with app_obj.app_context():
            for email in emails:
                try:
                    msg = Message(subject=asunto, recipients=[email])
                    msg.body = cuerpo
                    mail.send(msg)
                except Exception as e:
                    app_obj.logger.error(f"Error enviando mensaje masivo a {email}: {e}")

    threading.Thread(target=enviar_todos, daemon=True).start()

    registrar_auditoria(
        f"Envió mensaje masivo a {len(emails)} clientes: {asunto}",
        "Clientes", "clientes", None,
        f"Filtro: {destinatarios}"
    )

    return redirect(f"/admin/mensajes?enviado={len(emails)}")


# ── Papelera de reciclaje ────────────────────────────────────────────────────

@admin_bp.route("/papelera")
def papelera():
    if session.get("rol") != "admin":
        return redirect("/admin?access_error=Acceso+restringido+a+administradores")
    conexion = conectar()
    cursor = conexion.cursor()

    cursor.execute("""
        SELECT * FROM camioneros
        WHERE deleted_at IS NOT NULL
        ORDER BY deleted_at DESC
    """)
    camioneros_eliminados = cursor.fetchall()

    cursor.execute("""
        SELECT * FROM clientes
        WHERE deleted_at IS NOT NULL
        ORDER BY deleted_at DESC
    """)
    clientes_eliminados = cursor.fetchall()

    cursor.execute("""
        SELECT * FROM viajes
        WHERE deleted_at IS NOT NULL
        ORDER BY deleted_at DESC
    """)
    viajes_eliminados = cursor.fetchall()

    cursor.execute("""
        SELECT * FROM solicitudes_eliminacion
        WHERE estado = 'Pendiente'
        ORDER BY fecha_solicitud DESC
    """)
    solicitudes = cursor.fetchall()

    conexion.close()
    return render_template("admin/papelera.html",
                           camioneros_eliminados=camioneros_eliminados,
                           clientes_eliminados=clientes_eliminados,
                           viajes_eliminados=viajes_eliminados,
                           solicitudes=solicitudes)


@admin_bp.route("/papelera/<entidad>/<int:id>/restaurar", methods=["POST"])
def restaurar_registro(entidad, id):
    if session.get("rol") != "admin":
        return redirect("/admin?access_error=Acceso+restringido+a+administradores")

    tablas = {"camionero": "camioneros", "cliente": "clientes", "viaje": "viajes",
              "camioneros": "camioneros", "clientes": "clientes", "viajes": "viajes"}
    tabla = tablas.get(entidad)
    if not tabla:
        return redirect("/admin/papelera")

    conexion = conectar()
    cursor = conexion.cursor()
    cursor.execute(f"UPDATE {tabla} SET deleted_at = NULL, deleted_by = NULL WHERE id = {ph()}", (id,))
    conexion.commit()
    conexion.close()
    registrar_auditoria("restauró de papelera", tabla, entidad, id)
    return redirect("/admin/papelera")


# ── Solicitudes de eliminación ───────────────────────────────────────────────

@admin_bp.route("/solicitudes-eliminacion/<int:id>/aprobar", methods=["POST"])
def aprobar_eliminacion(id):
    if session.get("rol") != "admin":
        return redirect("/admin?access_error=Acceso+restringido+a+administradores")
    conexion = conectar()
    cursor = conexion.cursor()
    cursor.execute(f"SELECT * FROM solicitudes_eliminacion WHERE id = {ph()}", (id,))
    sol = cursor.fetchone()
    if sol and sol["estado"] == "Pendiente":
        tablas = {"camionero": "camioneros", "cliente": "clientes", "viaje": "viajes"}
        tabla = tablas.get(sol["entidad"])
        if tabla:
            cursor.execute(f"""
                UPDATE {tabla} SET deleted_at = CURRENT_TIMESTAMP, deleted_by = ?
                WHERE id = ?
            """, (session.get("usuario"), sol["entidad_id"]))
        cursor.execute(f"""
            UPDATE solicitudes_eliminacion
            SET estado = 'Aprobada', revisado_por = {ph()}, fecha_revision = CURRENT_TIMESTAMP
            WHERE id = {ph()}
        """, (session.get("usuario"), id))
        conexion.commit()
        registrar_auditoria("aprobó eliminación", sol["entidad"] + "s", sol["entidad"], sol["entidad_id"],
                            f"Solicitado por {sol['solicitado_por']}")
    conexion.close()
    return redirect("/admin")


@admin_bp.route("/solicitudes-eliminacion/<int:id>/rechazar", methods=["POST"])
def rechazar_eliminacion(id):
    if session.get("rol") != "admin":
        return redirect("/admin?access_error=Acceso+restringido+a+administradores")
    conexion = conectar()
    cursor = conexion.cursor()
    cursor.execute(f"""
        UPDATE solicitudes_eliminacion
        SET estado = 'Rechazada', revisado_por = {ph()}, fecha_revision = CURRENT_TIMESTAMP
        WHERE id = {ph()} AND estado = 'Pendiente'
    """, (session.get("usuario"), id))
    conexion.commit()
    conexion.close()
    return redirect("/admin")
